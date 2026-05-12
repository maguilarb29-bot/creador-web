"""
Casa Pignatelli — Servidor de inventario y ventas
Arrancar: python server.py
URL:      http://localhost:8080
"""
from flask import Flask, jsonify, request, send_from_directory, abort
import json, os, re, uuid, threading, time
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).parent
DATA = BASE / "data"
CATALOGO_FILE      = DATA / "solaris_catalogo.json"
HEREDEROS_FILE     = DATA / "reservas_excel_entregable_2026-04-13.json"
TRANSACCIONES_FILE = DATA / "transacciones.json"
ESTADOS_FILE       = DATA / "estados.json"
CONTADORES_FILE    = DATA / "contadores.json"
EXCEL_VENTAS       = DATA / "Registro_Ventas_Reservas.xlsx"
ENV_PATHS          = [BASE.parent / "pignatelli-app" / ".env.local", BASE.parent / ".env.local"]

app = Flask(__name__, static_folder=str(BASE))

# ── helpers ──────────────────────────────────────────────────
def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_env_file():
    env = {}
    for path in ENV_PATHS:
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            v = v.strip()
            if v.startswith('"') and v.endswith('"'):
                v = v[1:-1]
            env[k.strip()] = v
    env.update({k: v for k, v in os.environ.items() if k.startswith("GOOGLE_")})
    return env

def siguiente_factura():
    c = load_json(CONTADORES_FILE) if CONTADORES_FILE.exists() else {"ultimaFactura": 9}
    c["ultimaFactura"] = c.get("ultimaFactura", 9) + 1
    save_json(CONTADORES_FILE, c)
    return f"FAC-{c['ultimaFactura']:04d}"

def fmt_usd(value):
    if value in (None, ""):
        return ""
    try:
        n = float(value)
    except Exception:
        return str(value)
    return f"${int(n)}" if n.is_integer() else f"${n:.2f}"

def fmt_fecha(timestamp):
    try:
        return datetime.fromisoformat(timestamp).strftime("%d/%m/%Y")
    except Exception:
        return datetime.now().strftime("%d/%m/%Y")

def item_codes(items):
    return [str(i.get("codigoItem", "")).strip() for i in items if i.get("codigoItem")]

def load_estados():
    if not ESTADOS_FILE.exists():
        return {}
    try:
        return load_json(ESTADOS_FILE)
    except Exception:
        return {}

def save_estado(codigo, estado, reservadoPara="", confirmadoHeredero=False):
    estados = load_estados()
    estados[codigo] = {
        "estado": estado,
        "reservadoPara": reservadoPara,
        "confirmadoHeredero": confirmadoHeredero,
    }
    save_json(ESTADOS_FILE, estados)

def catalogo_con_estados():
    catalogo = load_json(CATALOGO_FILE)
    estados = load_estados()
    for item in catalogo:
        e = estados.get(item["codigoItem"])
        if e:
            item["estado"] = e["estado"]
            item["reservadoPara"] = e.get("reservadoPara", "")
            item["confirmadoHeredero"] = e.get("confirmadoHeredero", False)
    return catalogo

def load_transacciones():
    if not TRANSACCIONES_FILE.exists():
        return {"transacciones": []}
    return load_json(TRANSACCIONES_FILE)

def header_cell(ws, row, col, value, bg="1a1209", fg="c9a227", bold=True, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="333333")
    c.border = Border(thin, thin, thin, thin)
    return c

def auto_width(ws, extra=3, max_w=60):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        first = col[0]
        if isinstance(first, MergedCell):
            continue
        w = max((len(str(c.value)) for c in col if c.value and not isinstance(c, MergedCell)), default=8)
        ws.column_dimensions[first.column_letter].width = min(w + extra, max_w)

# ── Excel export ─────────────────────────────────────────────
def rebuild_excel(transacciones):
    wb = Workbook()

    # ── Hoja Transacciones ──
    ws = wb.active
    ws.title = "Transacciones"
    ws.row_dimensions[1].height = 32

    cols_tx = ["ID Tiquete", "Fecha/Hora", "Tipo", "Heredero",
               "N° Artículos", "Est. Mín USD", "Est. Máx USD",
               "Precio Acordado USD", "Estado", "Notas"]
    for i, h in enumerate(cols_tx, 1):
        header_cell(ws, 1, i, h)

    for r, tx in enumerate(transacciones, 2):
        data_row = [
            tx["id"],
            tx["timestamp"][:16].replace("T", " "),
            tx["tipo"].upper(),
            tx["heredero"],
            len(tx.get("items", [])),
            tx.get("totalEstMinUSD", 0),
            tx.get("totalEstMaxUSD", 0),
            tx.get("totalAcordadoUSD") or "",
            tx["estado"].upper(),
            tx.get("notas", ""),
        ]
        row_color = "f0fff0" if tx["estado"] == "confirmado" else "fff0f0"
        for c_idx, val in enumerate(data_row, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=row_color)
            thin = Side(border_style="thin", color="cccccc")
            cell.border = Border(thin, thin, thin, thin)
            cell.alignment = Alignment(vertical="center")

    auto_width(ws)

    # ── Hoja Detalle Artículos ──
    ws2 = wb.create_sheet("Detalle Artículos")
    ws2.row_dimensions[1].height = 32
    cols_det = ["ID Tiquete", "Heredero", "Tipo", "Código", "Artículo",
                "Categoría", "Est. Mín USD", "Est. Máx USD", "Precio Acordado", "Ref Sotheby's"]
    for i, h in enumerate(cols_det, 1):
        header_cell(ws2, 1, i, h)

    r2 = 2
    for tx in transacciones:
        for item in tx.get("items", []):
            row_color = "f0fff0" if tx["estado"] == "confirmado" else "fff0f0"
            vals = [
                tx["id"], tx["heredero"], tx["tipo"].upper(),
                item.get("codigoItem", ""), item.get("nombreES", ""),
                item.get("categoria", ""),
                item.get("estimMin", 0), item.get("estimMax", 0),
                item.get("precioAcordado") or "",
                item.get("refSothebys", ""),
            ]
            for c_idx, val in enumerate(vals, 1):
                cell = ws2.cell(row=r2, column=c_idx, value=val)
                cell.fill = PatternFill("solid", fgColor=row_color)
                thin = Side(border_style="thin", color="cccccc")
                cell.border = Border(thin, thin, thin, thin)
                cell.alignment = Alignment(vertical="center")
            r2 += 1

    auto_width(ws2)

    # ── Hoja Resumen ──
    ws3 = wb.create_sheet("Resumen")
    ws3.merge_cells("A1:D1")
    c = ws3["A1"]
    c.value = "Casa Pignatelli — Resumen de Ventas y Reservas"
    c.font = Font(bold=True, size=14, color="c9a227")
    c.fill = PatternFill("solid", fgColor="1a1209")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 40

    header_cell(ws3, 2, 1, "Heredero")
    header_cell(ws3, 2, 2, "Reservas")
    header_cell(ws3, 2, 3, "Ventas")
    header_cell(ws3, 2, 4, "Total Est. Máx USD")

    heirs = {}
    for tx in transacciones:
        if tx["estado"] != "confirmado":
            continue
        h = tx["heredero"]
        if h not in heirs:
            heirs[h] = {"reservas": 0, "ventas": 0, "max": 0}
        if tx["tipo"] == "reserva":
            heirs[h]["reservas"] += 1
        else:
            heirs[h]["ventas"] += 1
        heirs[h]["max"] += tx.get("totalEstMaxUSD", 0)

    for r3, (heir, v) in enumerate(sorted(heirs.items()), 3):
        ws3.cell(row=r3, column=1, value=heir)
        ws3.cell(row=r3, column=2, value=v["reservas"])
        ws3.cell(row=r3, column=3, value=v["ventas"])
        ws3.cell(row=r3, column=4, value=v["max"])

    auto_width(ws3)
    wb.save(EXCEL_VENTAS)
    return str(EXCEL_VENTAS.name)

# ── helpers de estado ────────────────────────────────────────
def parse_estimate(s):
    if not s: return {"min": 0, "max": 0}
    import re
    nums = re.findall(r"\d[\d,]*", s.replace("$", ""))
    if len(nums) >= 2:
        return {"min": int(nums[0].replace(",","")), "max": int(nums[1].replace(",",""))}
    return {"min": 0, "max": 0}

def active_transactions_for_codes(transacciones, codes):
    wanted = set(codes)
    found = []
    for tx in transacciones:
        if tx.get("estado") != "confirmado":
            continue
        tx_codes = set(item_codes(tx.get("items", [])))
        if tx_codes & wanted:
            found.append(tx)
    return found

def build_tx(body, tipo, heredero, items, notas):
    total_min = sum(i.get("estimMin", 0) or 0 for i in items)
    total_max = sum(i.get("estimMax", 0) or 0 for i in items)
    total_acordado = body.get("totalAcordadoUSD")
    if total_acordado in ("", None):
        total_acordado = sum(i.get("precioAcordado", 0) or 0 for i in items) or None
    else:
        total_acordado = float(total_acordado)

    tx = {
        "id": f"TX-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{str(uuid.uuid4())[:4].upper()}",
        "timestamp": datetime.now().isoformat(),
        "tipo": tipo,
        "heredero": heredero,
        "items": items,
        "totalEstMinUSD": total_min,
        "totalEstMaxUSD": total_max,
        "totalAcordadoUSD": total_acordado,
        "notas": notas,
        "estado": "confirmado",
    }
    if tipo == "venta":
        tx["numeroFactura"] = siguiente_factura()
        for item in tx["items"]:
            if not item.get("precioAcordado") and len(tx["items"]) == 1:
                item["precioAcordado"] = total_acordado
    return tx

# ── Auto-sync desde Google Sheets ───────────────────────────
_catalogo_cache: dict = {"items": None, "ts": 0.0}
_catalogo_lock  = threading.Lock()
CACHE_TTL       = 300  # segundos (5 minutos)


def _cell(row: list, idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    return str(row[idx]).strip()


def _parse_precio(raw) -> "float | None":
    if raw is None:
        return None
    s = re.sub(r"[$€\s]", "", str(raw).strip())
    if not s or s in ("—", "-"):
        return None
    try:
        val = float(s.replace(",", ""))
        return val if val > 0 else None
    except Exception:
        return None


def _col_idx(header: list, fragment: str) -> int:
    frag = fragment.lower()
    for i, h in enumerate(header):
        if frag in str(h).lower():
            return i
    return -1


def _sync_catalogo_desde_sheet():
    """Lee INVENTARIO_MAESTRO y actualiza _catalogo_cache (y el JSON en disco)."""
    svc, sheet_id, error = get_sheets_service()
    if error:
        return

    rows = sheet_values(svc, sheet_id, "INVENTARIO_MAESTRO!A1:N600")
    if not rows:
        return

    header = [str(h).strip() for h in rows[0]]
    ci = {
        "cod":       _col_idx(header, "artículo") if _col_idx(header, "artículo") >= 0 else _col_idx(header, "articulo"),
        "nombre":    _col_idx(header, "nombre"),
        "cat":       _col_idx(header, "categoría") if _col_idx(header, "categoría") >= 0 else _col_idx(header, "categoria"),
        "precio":    _col_idx(header, "precio usd"),
        "estado":    _col_idx(header, "estado"),
        "comprador": _col_idx(header, "reservado"),
        "est_soth":  _col_idx(header, "estimación") if _col_idx(header, "estimación") >= 0 else _col_idx(header, "estimacion"),
        "ref_soth":  _col_idx(header, "ref sotheby"),
        "pag_soth":  _col_idx(header, "página") if _col_idx(header, "página") >= 0 else _col_idx(header, "pagina"),
        "notas":     _col_idx(header, "notas"),
    }

    sheet_map: dict = {}
    for row in rows[1:]:
        cod = _cell(row, ci["cod"])
        if not cod:
            continue
        est_soth = _cell(row, ci["est_soth"])
        ref_soth = _cell(row, ci["ref_soth"])
        pag_soth = _cell(row, ci["pag_soth"])
        sheet_map[cod] = {
            "codigoItem":         cod,
            "nombreES":           _cell(row, ci["nombre"]),
            "categoria":          _cell(row, ci["cat"]),
            "precioUSD":          _parse_precio(_cell(row, ci["precio"])),
            "estado":             _cell(row, ci["estado"]) or "Disponible",
            "reservadoPara":      _cell(row, ci["comprador"]),
            "tieneSothebys":      bool(est_soth and est_soth not in ("—", "-")) or
                                  bool(ref_soth and ref_soth not in ("—", "-")),
            "estimacionSothebys": est_soth if est_soth not in ("—", "-") else "",
            "refSothebys":        ref_soth if ref_soth not in ("—", "-") else "",
            "paginaSothebys":     pag_soth if pag_soth not in ("—", "-") else "",
            "notas":              _cell(row, ci["notas"]),
        }

    old_map: dict = {}
    if CATALOGO_FILE.exists():
        try:
            old_map = {i["codigoItem"]: i for i in load_json(CATALOGO_FILE)
                       if isinstance(i, dict) and i.get("codigoItem")}
        except Exception:
            pass

    new_items = []
    for cod, sd in sheet_map.items():
        if cod in old_map:
            item = dict(old_map[cod])
            for k in ("nombreES", "categoria", "precioUSD", "estado", "reservadoPara",
                      "tieneSothebys", "estimacionSothebys", "refSothebys", "paginaSothebys"):
                item[k] = sd[k]
            if sd["notas"]:
                item["notas"] = sd["notas"]
        else:
            item = {
                "codigoItem": cod, "codigoPadre": "", "numItem": 0,
                "tipoEstructural": "ARTICULO", "fotos": [],
                "descripcionES": "", "descripcionSothebys": "", "cantidad": 1,
                **sd,
            }
        new_items.append(item)

    try:
        save_json(CATALOGO_FILE, new_items)
    except Exception:
        pass

    with _catalogo_lock:
        _catalogo_cache["items"] = new_items
        _catalogo_cache["ts"]    = time.time()


def _bg_sync_loop():
    time.sleep(10)  # espera arranque del servidor
    while True:
        try:
            _sync_catalogo_desde_sheet()
        except Exception:
            pass
        time.sleep(CACHE_TTL)


threading.Thread(target=_bg_sync_loop, daemon=True).start()


def get_sheets_service():
    env = load_env_file()
    sheet_id = env.get("GOOGLE_SHEETS_ID", "").strip()
    email = env.get("GOOGLE_SERVICE_ACCOUNT_EMAIL", "").strip()
    key = env.get("GOOGLE_PRIVATE_KEY", "").replace("\\n", "\n").strip()
    if not sheet_id or not email or not key:
        return None, None, "Credenciales de Google Sheets no configuradas en el servidor"
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except Exception as exc:
        return None, None, f"Librerías Google no instaladas: {exc}"
    info = {
        "type": "service_account",
        "project_id": "inventario-pignatelli",
        "private_key_id": "",
        "private_key": key,
        "client_email": email,
        "client_id": "",
        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
        "token_uri": "https://oauth2.googleapis.com/token",
        "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
        "client_x509_cert_url": "",
    }
    try:
        creds = service_account.Credentials.from_service_account_info(
            info, scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        return build("sheets", "v4", credentials=creds, cache_discovery=False), sheet_id, None
    except Exception as exc:
        return None, None, f"No se pudo iniciar Google Sheets: {exc}"

def sheet_values(svc, sheet_id, range_name):
    return svc.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range=range_name,
        valueRenderOption="UNFORMATTED_VALUE",
    ).execute().get("values", [])

def normalize_header(s):
    s = str(s or "").strip().lower()
    for a, b in {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n"}.items():
        s = s.replace(a, b)
    return re.sub(r"[^a-z0-9]+", " ", s).strip()

def header_index(headers, names):
    normalized = [normalize_header(h) for h in headers]
    for name in names:
        n = normalize_header(name)
        if n in normalized:
            return normalized.index(n)
    return None

def find_sheet_row(rows, code):
    target = str(code).strip()
    for idx, row in enumerate(rows[1:], start=2):
        if row and str(row[0]).strip() == target:
            return idx, row
    return None, None

def sync_sheet_for_transaction(tx):
    svc, sheet_id, error = get_sheets_service()
    if error:
        return {"ok": False, "error": error}

    try:
        inv = sheet_values(svc, sheet_id, "INVENTARIO_MAESTRO!A1:K1000")
        ventas = sheet_values(svc, sheet_id, "VENTAS!A1:H1000")
        reservas = sheet_values(svc, sheet_id, "RESERVAS!A1:F1200")
        inv_headers = inv[0] if inv else []
        idx_estado = header_index(inv_headers, ["Estado"])
        idx_comprador = header_index(inv_headers, ["Reservado / Comprador", "Comprador", "Reservado Para"])
        updates = []
        clears = []
        estado = "Vendido" if tx["tipo"] == "venta" else "Reservado"
        fecha = fmt_fecha(tx["timestamp"])
        nota_base = tx.get("notas", "")
        if tx.get("numeroFactura"):
            nota_base = f"{nota_base} | {tx['numeroFactura']}".strip(" | ")

        for item in tx.get("items", []):
            code = item.get("codigoItem", "")
            inv_row, _ = find_sheet_row(inv, code)
            if inv_row and idx_estado is not None:
                col = chr(ord("A") + idx_estado)
                updates.append({"range": f"INVENTARIO_MAESTRO!{col}{inv_row}", "values": [[estado]]})
            if inv_row and idx_comprador is not None:
                col = chr(ord("A") + idx_comprador)
                updates.append({"range": f"INVENTARIO_MAESTRO!{col}{inv_row}", "values": [[tx.get("heredero", "")]]})

            if tx["tipo"] == "venta":
                res_row, _ = find_sheet_row(reservas, code)
                if res_row:
                    clears.append(f"RESERVAS!A{res_row}:F{res_row}")

                ven_row, _ = find_sheet_row(ventas, code)
                row_values = [[
                    code,
                    item.get("nombreES", ""),
                    item.get("categoria", ""),
                    tx.get("heredero", ""),
                    fmt_usd(item.get("precioAcordado") or tx.get("totalAcordadoUSD")),
                    "",
                    fecha,
                    nota_base,
                ]]
                if ven_row:
                    updates.append({"range": f"VENTAS!A{ven_row}:H{ven_row}", "values": row_values})
                else:
                    svc.spreadsheets().values().append(
                        spreadsheetId=sheet_id,
                        range="VENTAS!A:H",
                        valueInputOption="USER_ENTERED",
                        insertDataOption="INSERT_ROWS",
                        body={"values": row_values},
                    ).execute()
            else:
                res_row, _ = find_sheet_row(reservas, code)
                row_values = [[
                    code,
                    item.get("nombreES", ""),
                    item.get("categoria", ""),
                    tx.get("heredero", ""),
                    fmt_usd(item.get("precioAcordado") or item.get("estimMin")),
                    nota_base,
                ]]
                if res_row:
                    updates.append({"range": f"RESERVAS!A{res_row}:F{res_row}", "values": row_values})
                else:
                    svc.spreadsheets().values().append(
                        spreadsheetId=sheet_id,
                        range="RESERVAS!A:F",
                        valueInputOption="USER_ENTERED",
                        insertDataOption="INSERT_ROWS",
                        body={"values": row_values},
                    ).execute()

        if clears:
            svc.spreadsheets().values().batchClear(
                spreadsheetId=sheet_id, body={"ranges": clears}
            ).execute()
        if updates:
            svc.spreadsheets().values().batchUpdate(
                spreadsheetId=sheet_id,
                body={"valueInputOption": "USER_ENTERED", "data": updates},
            ).execute()
        return {"ok": True}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}

def create_transaction(body):
    items = body.get("items", [])
    tipo = body.get("tipo", "reserva").strip().lower()
    heredero = body.get("heredero", "").strip()
    notas = body.get("notas", "").strip()
    if tipo not in ("reserva", "venta"):
        return None, (jsonify({"error": "Tipo inválido"}), 400)
    if not items:
        return None, (jsonify({"error": "Sin artículos"}), 400)
    if not heredero:
        return None, (jsonify({"error": "Debe indicar el heredero o comprador"}), 400)

    codes = item_codes(items)
    catalogo = {i["codigoItem"]: i for i in catalogo_con_estados()}
    missing = [c for c in codes if c not in catalogo]
    if missing:
        return None, (jsonify({"error": f"Artículo no encontrado: {', '.join(missing)}"}), 404)

    unavailable = []
    for code in codes:
        estado_actual = str(catalogo[code].get("estado", "Disponible")).lower()
        if tipo == "reserva" and estado_actual != "disponible":
            unavailable.append(code)
        if tipo == "venta" and estado_actual == "vendido":
            unavailable.append(code)
    if unavailable:
        return None, (jsonify({"error": f"Artículo no disponible: {', '.join(unavailable)}"}), 409)

    data = load_transacciones()
    conflicts = active_transactions_for_codes(data["transacciones"], codes)
    active_sales = [t for t in conflicts if t.get("tipo") == "venta"]
    if active_sales:
        sold = ", ".join(sorted(set(c for t in active_sales for c in item_codes(t.get("items", []))) & set(codes)))
        return None, (jsonify({"error": f"Ya vendido: {sold}"}), 409)
    if tipo == "reserva" and conflicts:
        reserved = ", ".join(sorted(set(c for t in conflicts for c in item_codes(t.get("items", []))) & set(codes)))
        return None, (jsonify({"error": f"Ya reservado: {reserved}"}), 409)

    tx = build_tx(body, tipo, heredero, items, notas)
    if tipo == "venta" and not tx.get("totalAcordadoUSD"):
        return None, (jsonify({"error": "Toda venta debe tener precio acordado"}), 400)

    if tipo == "venta":
        for old in conflicts:
            if old.get("tipo") == "reserva":
                old["estado"] = "convertido"
                old["convertidoEn"] = tx["timestamp"]
                old["ventaId"] = tx["id"]

    data["transacciones"].append(tx)
    save_json(TRANSACCIONES_FILE, data)

    nuevo_estado = "Reservado" if tipo == "reserva" else "Vendido"
    for code in codes:
        save_estado(code, nuevo_estado, heredero)
    rebuild_excel(data["transacciones"])
    tx["sheetSync"] = sync_sheet_for_transaction(tx)
    return tx, None

# ── Rutas API ────────────────────────────────────────────────
@app.route("/api/catalogo")
def api_catalogo():
    ahora = time.time()
    with _catalogo_lock:
        stale = _catalogo_cache["items"] is None or (ahora - _catalogo_cache["ts"]) > CACHE_TTL
        items = _catalogo_cache["items"]

    if stale:
        try:
            _sync_catalogo_desde_sheet()
            with _catalogo_lock:
                items = _catalogo_cache["items"]
        except Exception:
            pass

    return jsonify(items or catalogo_con_estados())

@app.route("/api/sync", methods=["POST"])
def api_sync():
    try:
        _sync_catalogo_desde_sheet()
        with _catalogo_lock:
            ts = _catalogo_cache["ts"]
            n  = len(_catalogo_cache["items"] or [])
        return jsonify({"ok": True, "items": n, "ts": ts})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/herederos")
def api_herederos():
    return jsonify(load_json(HEREDEROS_FILE))

@app.route("/api/transacciones")
def api_transacciones():
    return jsonify(load_transacciones())

@app.route("/api/transaccion", methods=["POST"])
def api_crear_transaccion():
    body = request.get_json(force=True)
    tx, error = create_transaction(body)
    if error:
        return error
    return jsonify({"success": True, "transaccion": tx})
    items = body.get("items", [])
    tipo  = body.get("tipo", "reserva")
    heredero = body.get("heredero", "").strip()
    notas    = body.get("notas", "").strip()

    if not items:
        return jsonify({"error": "Sin artículos"}), 400
    if not heredero:
        return jsonify({"error": "Debe indicar el heredero o comprador"}), 400

    tx_id = f"TX-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{str(uuid.uuid4())[:4].upper()}"

    total_min = sum(i.get("estimMin", 0) or 0 for i in items)
    total_max = sum(i.get("estimMax", 0) or 0 for i in items)
    total_acordado = sum(i.get("precioAcordado", 0) or 0 for i in items) or None

    tx = {
        "id": tx_id,
        "timestamp": datetime.now().isoformat(),
        "tipo": tipo,
        "heredero": heredero,
        "items": items,
        "totalEstMinUSD": total_min,
        "totalEstMaxUSD": total_max,
        "totalAcordadoUSD": total_acordado,
        "notas": notas,
        "estado": "confirmado",
    }

    data = load_transacciones()
    data["transacciones"].append(tx)
    save_json(TRANSACCIONES_FILE, data)

    # Actualizar estado
    nuevo_estado = "Reservado" if tipo == "reserva" else "Vendido"
    for i in items:
        save_estado(i["codigoItem"], nuevo_estado, heredero)
    rebuild_excel(data["transacciones"])

    return jsonify({"success": True, "transaccion": tx})

@app.route("/api/reservar-item", methods=["POST"])
def api_reservar_item():
    body = request.get_json(force=True)
    codigo   = body.get("codigoItem", "").strip()
    heredero = body.get("heredero", "").strip()
    if not codigo or not heredero:
        return jsonify({"error": "Faltan datos"}), 400

    catalogo = catalogo_con_estados()
    item = next((i for i in catalogo if i["codigoItem"] == codigo), None)
    if not item:
        return jsonify({"error": "Artículo no encontrado"}), 404

    est = parse_estimate(item.get("estimacionSothebys", ""))
    tx, error = create_transaction({
        "tipo": "reserva",
        "heredero": heredero,
        "notas": "Reserva directa Sotheby's",
        "items": [{
            "codigoItem": codigo,
            "nombreES": item.get("nombreES", ""),
            "categoria": item.get("categoria", ""),
            "estimMin": est["min"],
            "estimMax": est["max"],
            "estimacionSothebys": item.get("estimacionSothebys", ""),
            "refSothebys": item.get("refSothebys", ""),
            "fotos": item.get("fotos", []),
        }],
    })
    if error:
        return error
    return jsonify({"success": True, "transaccion": tx})

    save_estado(codigo, "Reservado", heredero, False)

    # Registrar transacción
    est = parse_estimate(item.get("estimacionSothebys", ""))
    data = load_transacciones()
    tx_id = f"TX-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{str(uuid.uuid4())[:4].upper()}"
    tx = {
        "id": tx_id,
        "timestamp": datetime.now().isoformat(),
        "tipo": "reserva",
        "heredero": heredero,
        "items": [{
            "codigoItem":        codigo,
            "nombreES":          item.get("nombreES", ""),
            "categoria":         item.get("categoria", ""),
            "estimMin":          est["min"],
            "estimMax":          est["max"],
            "estimacionSothebys":item.get("estimacionSothebys", ""),
            "refSothebys":       item.get("refSothebys", ""),
            "fotos":             item.get("fotos", []),
        }],
        "totalEstMinUSD": est["min"],
        "totalEstMaxUSD": est["max"],
        "notas": "Reserva directa Sotheby's",
        "estado": "confirmado",
    }
    data["transacciones"].append(tx)
    save_json(TRANSACCIONES_FILE, data)
    rebuild_excel(data["transacciones"])
    return jsonify({"success": True, "transaccion": tx})

@app.route("/api/confirmar-item", methods=["POST"])
def api_confirmar_item():
    body = request.get_json(force=True)
    codigo = body.get("codigoItem", "").strip()
    estados = load_estados()
    e = estados.get(codigo, {})
    save_estado(codigo, e.get("estado", "Reservado"), e.get("reservadoPara", ""), True)
    return jsonify({"success": True})

@app.route("/api/liberar-item", methods=["POST"])
def api_liberar_item():
    body = request.get_json(force=True)
    codigo = body.get("codigoItem", "").strip()
    catalogo = catalogo_con_estados()
    item = next((i for i in catalogo if i["codigoItem"] == codigo), None)
    if not item:
        return jsonify({"error": "No encontrado"}), 404

    save_estado(codigo, "Disponible", "", False)

    # Cancelar transacciones activas de este artículo
    data = load_transacciones()
    for tx in data["transacciones"]:
        if tx["estado"] == "confirmado" and any(
            i.get("codigoItem") == codigo for i in tx.get("items", [])
        ):
            tx["estado"]      = "cancelado"
            tx["canceladoEn"] = datetime.now().isoformat()
    save_json(TRANSACCIONES_FILE, data)
    rebuild_excel(data["transacciones"])
    return jsonify({"success": True})

@app.route("/api/transaccion/<tx_id>/cancelar", methods=["POST"])
def api_cancelar(tx_id):
    data = load_transacciones()
    tx = next((t for t in data["transacciones"] if t["id"] == tx_id), None)
    if not tx:
        return jsonify({"error": "No encontrado"}), 404

    tx["estado"] = "cancelado"
    tx["canceladoEn"] = datetime.now().isoformat()

    # Revertir estado
    for i in tx.get("items", []):
        save_estado(i["codigoItem"], "Disponible", "")

    save_json(TRANSACCIONES_FILE, data)
    rebuild_excel(data["transacciones"])
    return jsonify({"success": True})

@app.route("/api/transaccion/<tx_id>/numero-factura", methods=["POST"])
def api_numero_factura(tx_id):
    data = load_transacciones()
    tx = next((t for t in data["transacciones"] if t["id"] == tx_id), None)
    if not tx:
        return jsonify({"error": "No encontrado"}), 404
    if not tx.get("numeroFactura"):
        tx["numeroFactura"] = siguiente_factura()
        save_json(TRANSACCIONES_FILE, data)
    return jsonify({"numeroFactura": tx["numeroFactura"]})

@app.route("/api/orden", methods=["POST"])
def api_crear_orden():
    data = request.json or {}
    c = load_json(CONTADORES_FILE) if CONTADORES_FILE.exists() else {"ultimaFactura": 9, "ultimaOrden": 0}
    c["ultimaOrden"] = c.get("ultimaOrden", 0) + 1
    num = f"ORD-{c['ultimaOrden']:04d}"
    save_json(CONTADORES_FILE, c)
    ordenes_file = DATA / "ordenes.json"
    ordenes = load_json(ordenes_file) if ordenes_file.exists() else {"ordenes": []}
    ordenes["ordenes"].append({
        "id": num,
        "timestamp": datetime.now().isoformat(),
        "cliente": data.get("cliente", ""),
        "telefono": data.get("telefono", ""),
        "email": data.get("email", ""),
        "items": data.get("items", []),
        "totalUSD": data.get("totalUSD", 0),
        "estado": "pendiente"
    })
    save_json(ordenes_file, ordenes)
    return jsonify({"success": True, "numeroOrden": num})

@app.route("/api/excel/descargar")
def api_descargar_excel():
    from flask import send_file
    if not EXCEL_VENTAS.exists():
        data = load_transacciones()
        rebuild_excel(data.get("transacciones", []))
    return send_file(str(EXCEL_VENTAS), as_attachment=True,
                     download_name="Pignatelli_Ventas_Reservas.xlsx")

# ── Archivos estáticos ───────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(str(BASE), "panel_herederos.html")

@app.route("/publico")
def publico():
    return send_from_directory(str(BASE), "solaris_catalogo.html")

@app.route("/images/<path:filename>")
def serve_images(filename):
    from flask import send_file
    import urllib.parse
    decoded = urllib.parse.unquote(filename)
    full_path = BASE / "images" / decoded
    if full_path.exists():
        return send_file(str(full_path))
    # Fallback: buscar solo por nombre de archivo en Todas las Fotos
    nombre = Path(decoded).name
    fallback = BASE / "images" / "fotos-Solaris-inventory" / "Todas las Fotos" / nombre
    if fallback.exists():
        return send_file(str(fallback))
    abort(404)

@app.route("/data/<path:filename>")
def serve_data(filename):
    return send_from_directory(str(BASE / "data"), filename)

@app.route("/<path:filename>")
def static_files(filename):
    full = BASE / filename
    if full.exists() and full.is_file():
        return send_from_directory(str(BASE), filename)
    abort(404)

if __name__ == "__main__":
    print("\nCasa Pignatelli - Servidor activo")
    print("  http://localhost:8090\n")
    app.run(host="0.0.0.0", port=8090, debug=False, use_reloader=False, threaded=True)
