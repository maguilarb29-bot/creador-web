from pathlib import Path
import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build

PROJECT_ROOT = Path(r'c:\Users\Alejandro\Documents\Proyecto Pignatelli')
ENV_PATH = PROJECT_ROOT / 'pignatelli-app' / '.env.local'

# -------- env + sheets client --------
def load_env(path):
    env={}
    for line in path.read_text(encoding='utf-8').splitlines():
        line=line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k,v=line.split('=',1)
        v=v.strip()
        if v.startswith('"') and v.endswith('"'):
            v=v[1:-1]
        env[k.strip()]=v
    return env

env=load_env(ENV_PATH)
info={
    'type':'service_account','project_id':'inventario-pignatelli','private_key_id':'',
    'private_key':env['GOOGLE_PRIVATE_KEY'].replace('\\n','\n').strip(),
    'client_email':env['GOOGLE_SERVICE_ACCOUNT_EMAIL'].strip(),'client_id':'',
    'auth_uri':'https://accounts.google.com/o/oauth2/auth','token_uri':'https://oauth2.googleapis.com/token',
    'auth_provider_x509_cert_url':'https://www.googleapis.com/oauth2/v1/certs','client_x509_cert_url':''
}
creds=service_account.Credentials.from_service_account_info(info, scopes=['https://www.googleapis.com/auth/spreadsheets'])
svc=build('sheets','v4',credentials=creds,cache_discovery=False)
sid=env['GOOGLE_SHEETS_ID'].strip()

# -------- check D11 in INVENTARIO_MAESTRO --------
vals=svc.spreadsheets().values().get(
    spreadsheetId=sid,
    range='INVENTARIO_MAESTRO!A1:N600',
    valueRenderOption='UNFORMATTED_VALUE'
).execute().get('values',[])
headers=vals[0]
rows=vals[1:]
idx={h:i for i,h in enumerate(headers)}

found=[]
for r in rows:
    code = r[idx['Artículo']] if idx['Artículo'] < len(r) else ''
    if str(code).strip()=='D11':
        found.append(r)

print('D11 en INVENTARIO_MAESTRO:', 'SI' if found else 'NO')
if found:
    r=found[0]
    def g(col):
        i=idx[col]
        return r[i] if i < len(r) else ''
    print('  Nombre:', g('Nombre'))
    print('  Precio USD:', g('Precio USD'))
    print('  Estado:', g('Estado'))
    print('  Cliente:', g('Cliente'))

# -------- check D11 in RESERVAS --------
res=svc.spreadsheets().values().get(
    spreadsheetId=sid,
    range='RESERVAS!A1:G300',
    valueRenderOption='UNFORMATTED_VALUE'
).execute().get('values',[])
res_headers=res[0] if res else []
res_rows=res[1:] if len(res)>1 else []
print('D11 en RESERVAS:', any(str(r[0]).strip()=='11' for r in res_rows if r))
if res_rows:
    for r in res_rows:
        if r and str(r[0]).strip()=='11':
            print('  Fila RESERVAS #11:', r)

# -------- verify boss price source for item #11 from both xlsx --------
files=[
    PROJECT_ROOT/'docs'/'Condominio Solaris-Belongings-01032026 (3).xlsx',
    PROJECT_ROOT/'docs'/'External-Condominio Solaris-Belongings.xlsx'
]
for f in files:
    wb=openpyxl.load_workbook(f, data_only=True, read_only=True)
    ws=wb.active
    row11=None
    for row in ws.iter_rows(min_row=6, max_row=400, values_only=True):
        item_num=row[1]
        if item_num==11:
            row11=row
            break
    print(f'Archivo: {f.name}')
    if row11:
        print('  item#11 desc:', row11[2])
        print('  notes:', row11[4])
        # ask_usd col index 7 in first file; second file only has Price Offer col5
        print('  precio referencia cols:', row11[5:10])
    else:
        print('  item#11 no encontrado')
    wb.close()

# -------- quick global price sanity against main boss file (ask USD) --------
main_file=PROJECT_ROOT/'docs'/'Condominio Solaris-Belongings-01032026 (3).xlsx'
wb=openpyxl.load_workbook(main_file, data_only=True, read_only=True)
ws=wb.active
price_map={}
for row in ws.iter_rows(min_row=6, max_row=400, values_only=True):
    num=row[1]
    if num is None:
        continue
    ask_usd=row[7]
    if ask_usd is not None and ask_usd!=0:
        price_map[int(num)] = float(ask_usd)
wb.close()

# compare against maestro by num extracted from code
mismatch=[]
missing=[]
for r in rows:
    code = r[idx['Artículo']] if idx['Artículo'] < len(r) else ''
    if not code:
        continue
    import re
    m=re.match(r'^[A-Z]+(\d+)', str(code))
    if not m:
        continue
    num=int(m.group(1))
    if num in price_map:
        p_sheet = r[idx['Precio USD']] if idx['Precio USD'] < len(r) else ''
        try:
            p_sheet=float(p_sheet)
        except:
            p_sheet=None
        if p_sheet is None:
            missing.append((code,num,price_map[num]))
        elif abs(p_sheet-price_map[num])>0.001:
            mismatch.append((code,num,p_sheet,price_map[num]))

print('--- resumen cruce precios (archivo principal de jefes) ---')
print('  codigos con precio esperado pero vacío en maestro:', len(missing))
print('  codigos con precio distinto:', len(mismatch))
print('  muestra missing:', missing[:8])
print('  muestra mismatch:', mismatch[:8])
