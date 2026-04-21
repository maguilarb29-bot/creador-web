from __future__ import annotations

import io
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

PROJECT_ROOT = Path("C:/Users/Alejandro/Documents/Proyecto Pignatelli")
SERVICE_ACCOUNT = PROJECT_ROOT / "pignatelli-app" / "lib" / "service-account.json"
MASTER_XLSX = PROJECT_ROOT / "docs" / "Inventario_Maestro_Solaris_Pignatelli_2026-04-14.xlsx"
CLIENTS_XLSX = PROJECT_ROOT / "docs" / "Ventas_Reservas_Clientes_2026-04-17.xlsx"
SHEET_ID = "1X6l9MeiFgXKRmVyRYBwsMnjwkWz971BsrBK3lbWJ-u4"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
EXCLUDED_CODES = {"216A", "299A"}


def norm_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_currency(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text in {"—", "â€”"} or text.lower() == "sin precio":
        return ""
    text = text.replace("CRC", "").replace("$", "").replace(",", "").strip()
    try:
        num = float(text)
    except ValueError:
        return ""
    return int(num) if num.is_integer() else num


def row_is_data(row) -> bool:
    return isinstance(row[0], (int, float)) and norm_str(row[1]) != ""


def fmt_usd(value: float) -> str:
    return f"${value:,.2f}"


def load_inventory_rows():
    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    ws = wb["Inventario Completo"]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row_is_data(row):
            continue
        code = norm_str(row[1])
        if code in EXCLUDED_CODES:
            continue
        rows.append(
            {
                "code": code,
                "name": norm_str(row[2]),
                "category": norm_str(row[3]),
                "valuation": norm_str(row[4]),
                "price": parse_currency(row[5]),
                "state": norm_str(row[6]) or "Disponible",
                "holder": norm_str(row[7]).replace("â€”", "").replace("—", "").strip(),
            }
        )
    wb.close()
    return rows


def load_sales_rows():
    wb = openpyxl.load_workbook(CLIENTS_XLSX, read_only=True, data_only=True)
    ws = wb["Ventas Clientes"]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row_is_data(row):
            continue
        code = norm_str(row[1])
        if code in EXCLUDED_CODES:
            continue
        rows.append(
            {
                "code": code,
                "buyer": norm_str(row[3]),
                "category": norm_str(row[4]),
                "price_usd": parse_currency(row[5]),
                "price_crc": norm_str(row[6]),
            }
        )
    wb.close()
    return rows


def load_client_reservation_rows():
    wb = openpyxl.load_workbook(CLIENTS_XLSX, read_only=True, data_only=True)
    ws = wb["Reservas Clientes"]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row_is_data(row):
            continue
        code = norm_str(row[1])
        if code in EXCLUDED_CODES:
            continue
        rows.append(
            {
                "code": code,
                "client": norm_str(row[3]),
                "category": norm_str(row[4]),
                "price_usd": parse_currency(row[5]),
            }
        )
    wb.close()
    return rows


def build_summary_rows(inventory_rows, sales_rows, client_res_rows):
    total = len(inventory_rows)
    state_counts = Counter(row["state"] for row in inventory_rows)
    category_counts = Counter(row["category"] for row in inventory_rows)
    with_sothebys = sum(1 for row in inventory_rows if row["valuation"] not in {"", "—", "â€”"})
    listed_total = sum(float(row["price"]) for row in inventory_rows if isinstance(row["price"], (int, float)))
    priced_count = sum(1 for row in inventory_rows if isinstance(row["price"], (int, float)))
    unpriced_count = total - priced_count
    sold_total = sum(float(row["price_usd"]) for row in sales_rows if isinstance(row["price_usd"], (int, float)))
    avg_sold = sold_total / len(sales_rows) if sales_rows else 0
    unique_buyers = len({row["buyer"] for row in sales_rows if row["buyer"]})
    client_reserved_codes = {row["code"] for row in client_res_rows}
    reserved_clients = sum(1 for row in inventory_rows if row["state"] == "Reservado" and row["code"] in client_reserved_codes)
    reserved_heirs = state_counts.get("Reservado", 0) - reserved_clients

    rows = [
        ["RESUMEN EJECUTIVO", "", "", "OPERACIÓN", "", "", "VALOR", "", "", "CATEGORÍAS", ""],
        ["Métrica", "Valor", "", "Métrica", "Valor", "", "Métrica", "Valor", "", "Categoría", "Cantidad"],
        ["Total de elementos", total, "", "Reservas totales", state_counts.get("Reservado", 0), "", "Valor listado USD", fmt_usd(listed_total), "", "", ""],
        ["Disponibles", state_counts.get("Disponible", 0), "", "Reservas de clientes", reserved_clients, "", "Valor vendido USD", fmt_usd(sold_total), "", "", ""],
        ["Reservados", state_counts.get("Reservado", 0), "", "Reservas de herederos", reserved_heirs, "", "Ticket promedio venta", fmt_usd(avg_sold), "", "", ""],
        ["Vendidos", state_counts.get("Vendido", 0), "", "Ventas totales", len(sales_rows), "", "Artículos con precio", priced_count, "", "", ""],
        ["Con Sotheby's", with_sothebys, "", "Compradores únicos", unique_buyers, "", "Artículos sin precio", unpriced_count, "", "", ""],
    ]

    for idx, (category, qty) in enumerate(sorted(category_counts.items(), key=lambda item: (-item[1], item[0])), start=2):
        while len(rows) <= idx:
            rows.append(["", "", "", "", "", "", "", "", "", "", ""])
        rows[idx][9] = category
        rows[idx][10] = qty
    return rows


def get_service():
    sa_info = json.loads(SERVICE_ACCOUNT.read_text(encoding="utf-8"))
    creds = service_account.Credentials.from_service_account_info(sa_info, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def get_sheet_id(svc, title: str) -> int:
    meta = svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    for sheet in meta["sheets"]:
        props = sheet["properties"]
        if props["title"] == title:
            return props["sheetId"]
    raise RuntimeError(f"No encontré la hoja {title}")


def format_summary_sheet(svc, sheet_id_num: int, row_count: int):
    requests = [
        {
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id_num, "gridProperties": {"frozenRowCount": 2}},
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id_num, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 11},
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.10, "green": 0.07, "blue": 0.03},
                        "textFormat": {"foregroundColor": {"red": 0.97, "green": 0.84, "blue": 0.45}, "bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat)",
            }
        },
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id_num, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": 11},
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.20, "green": 0.29, "blue": 0.38},
                        "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}, "bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat)",
            }
        },
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id_num, "startRowIndex": 2, "endRowIndex": row_count, "startColumnIndex": 7, "endColumnIndex": 8},
                "cell": {"userEnteredFormat": {"horizontalAlignment": "RIGHT"}},
                "fields": "userEnteredFormat.horizontalAlignment",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id_num, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 11},
                "properties": {"pixelSize": 120},
                "fields": "pixelSize",
            }
        },
    ]

    width_overrides = {
        0: 210,
        1: 120,
        3: 210,
        4: 120,
        6: 200,
        7: 140,
        9: 180,
        10: 100,
    }
    for col_idx, pixel_size in width_overrides.items():
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {"sheetId": sheet_id_num, "dimension": "COLUMNS", "startIndex": col_idx, "endIndex": col_idx + 1},
                    "properties": {"pixelSize": pixel_size},
                    "fields": "pixelSize",
                }
            }
        )

    svc.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body={"requests": requests}).execute()


def main():
    inventory_rows = load_inventory_rows()
    sales_rows = load_sales_rows()
    client_res_rows = load_client_reservation_rows()
    summary_rows = build_summary_rows(inventory_rows, sales_rows, client_res_rows)

    svc = get_service()
    svc.spreadsheets().values().clear(spreadsheetId=SHEET_ID, range="RESUMEN!A:Z", body={}).execute()
    svc.spreadsheets().values().update(
        spreadsheetId=SHEET_ID,
        range="RESUMEN!A1",
        valueInputOption="USER_ENTERED",
        body={"values": summary_rows},
    ).execute()
    format_summary_sheet(svc, get_sheet_id(svc, "RESUMEN"), len(summary_rows))

    print("RESUMEN actualizado")
    print(json.dumps({"rows": len(summary_rows), "inventoryCount": len(inventory_rows), "salesCount": len(sales_rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
