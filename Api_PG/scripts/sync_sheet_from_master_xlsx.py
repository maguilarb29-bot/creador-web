from __future__ import annotations

import io
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

PROJECT_ROOT = Path("C:/Users/Alejandro/Documents/Proyecto Pignatelli")
SERVICE_ACCOUNT = PROJECT_ROOT / "pignatelli-app" / "lib" / "service-account.json"
MASTER_XLSX = PROJECT_ROOT / "docs" / "Inventario_Maestro_Solaris_Pignatelli_2026-04-14.xlsx"
CLIENTS_XLSX = PROJECT_ROOT / "docs" / "Ventas_Reservas_Clientes_2026-04-17.xlsx"
CATALOGO_JSON = PROJECT_ROOT / "Api_PG" / "data" / "solaris_catalogo.json"
FOTOS_DIR = PROJECT_ROOT / "Api_PG" / "images" / "fotos-Solaris-inventory" / "Todas las Fotos"
LOG_PATH = PROJECT_ROOT / "docs" / "sheet_sync_log_2026-04-18.json"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
SHEET_ID = "1X6l9MeiFgXKRmVyRYBwsMnjwkWz971BsrBK3lbWJ-u4"
EXCLUDED_CODES = {"216A", "299A"}


def norm_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_currency(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text == "—" or text.lower() == "sin precio":
        return ""
    text = text.replace("CRC", "").replace("$", "").replace(",", "").strip()
    try:
        num = float(text)
    except ValueError:
        return value
    return int(num) if num.is_integer() else num


def row_is_data(row) -> bool:
    return isinstance(row[0], (int, float)) and norm_str(row[1]) != ""


def load_catalogo():
    data = json.loads(CATALOGO_JSON.read_text(encoding="utf-8"))
    by_code = {}
    for item in data:
        code = norm_str(item.get("codigoItem"))
        if code:
            by_code[code] = item
    return by_code


def load_photo_map():
    photo_map = {}
    for file in FOTOS_DIR.iterdir():
        if not file.is_file():
            continue
        name = file.name
        code = name.split("-", 1)[0].strip()
        if code and code not in photo_map:
            photo_map[code] = name
    return photo_map


def build_inventory_rows():
    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    ws = wb["Inventario Completo"]
    rows = []
    stats = Counter()
    missing_photos = []
    catalogo = load_catalogo()
    photo_map = load_photo_map()

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row_is_data(row):
            continue
        code = norm_str(row[1])
        if code in EXCLUDED_CODES:
            continue
        name = norm_str(row[2])
        category = norm_str(row[3])
        valuation = norm_str(row[4])
        price = parse_currency(row[5])
        state = norm_str(row[6]) or "Disponible"
        holder = norm_str(row[7]).replace("—", "").strip()
        notes = norm_str(row[8])

        source = catalogo.get(code, {})
        ref_s = norm_str(source.get("refSothebys"))
        page_s = norm_str(source.get("paginaSothebys"))
        estim_s = norm_str(source.get("estimacionSothebys")) or valuation
        photo = photo_map.get(code, "")

        if not photo:
            missing_photos.append(code)
            notes = (notes + " | " if notes else "") + "Pendiente visual: sin foto localizada en carpeta."

        rows.append([
            code,
            name,
            category,
            price,
            state,
            holder,
            ref_s,
            page_s,
            estim_s,
            notes,
        ])
        stats[state] += 1

    wb.close()
    return rows, stats, missing_photos


def build_sales_rows():
    wb = openpyxl.load_workbook(CLIENTS_XLSX, read_only=True, data_only=True)
    ws = wb["Ventas Clientes"]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row_is_data(row):
            continue
        code = norm_str(row[1])
        if code in EXCLUDED_CODES:
            continue
        rows.append([
            code,
            norm_str(row[2]),
            norm_str(row[4]),
            norm_str(row[3]),
            parse_currency(row[5]),
            norm_str(row[6]),
            norm_str(row[7]),
            "",
        ])
    wb.close()
    return rows


def build_client_reservation_rows():
    wb = openpyxl.load_workbook(CLIENTS_XLSX, read_only=True, data_only=True)
    ws = wb["Reservas Clientes"]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row_is_data(row):
            continue
        code = norm_str(row[1])
        if code in EXCLUDED_CODES:
            continue
        rows.append([
            code,
            norm_str(row[2]),
            norm_str(row[3]),
            norm_str(row[4]),
            parse_currency(row[5]),
        ])
    wb.close()
    return rows


def build_reservation_rows(inventory_rows):
    return [
        [
            row[0],  # codigo
            row[1],  # nombre
            row[2],  # categoria
            row[5],  # reservado/comprador
            row[3],  # precio usd
            row[9],  # notas
        ]
        for row in inventory_rows
        if row[4] == "Reservado"
    ]


def build_summary_rows(inventory_rows, sales_rows):
    total = len(inventory_rows)
    counts = Counter(row[4] for row in inventory_rows)
    with_sothebys = sum(1 for row in inventory_rows if row[6] or row[8])
    listed_total = sum(float(row[3]) for row in inventory_rows if isinstance(row[3], (int, float)))
    sold_total_usd = sum(float(row[4]) if False else 0 for _ in [])  # placeholder for shape consistency
    sold_total_usd = sum(float(row[4]) for row in [] if False)
    sold_total_usd = sum(float(r[4]) for r in sales_rows if isinstance(r[4], (int, float)))

    categories = Counter(row[2] for row in inventory_rows)

    rows = [
        ["RESUMEN EJECUTIVO", ""],
        ["Métrica", "Valor"],
        ["Total de elementos", total],
        ["Disponibles", counts.get("Disponible", 0)],
        ["Reservados", counts.get("Reservado", 0)],
        ["Vendidos", counts.get("Vendido", 0)],
        ["Con Sotheby's", with_sothebys],
        ["Valor listado USD", listed_total],
        ["Valor vendido USD", sold_total_usd],
        ["", ""],
        ["Categoría", "Cantidad"],
    ]
    for category, qty in sorted(categories.items()):
        rows.append([category, qty])
    return rows


def get_service():
    sa_info = json.loads(SERVICE_ACCOUNT.read_text(encoding="utf-8"))
    creds = service_account.Credentials.from_service_account_info(sa_info, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def ensure_sheet(svc, title: str):
    meta = svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    for sheet in meta["sheets"]:
        props = sheet["properties"]
        if props["title"] == title:
            return props["sheetId"]

    body = {"requests": [{"addSheet": {"properties": {"title": title}}}]}
    svc.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body=body).execute()
    meta = svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    for sheet in meta["sheets"]:
        props = sheet["properties"]
        if props["title"] == title:
            return props["sheetId"]
    raise RuntimeError(f"No pude crear la hoja {title}")


def write_sheet(svc, title: str, rows):
    svc.spreadsheets().values().clear(
        spreadsheetId=SHEET_ID,
        range=f"{title}!A:Z",
        body={},
    ).execute()

    svc.spreadsheets().values().update(
        spreadsheetId=SHEET_ID,
        range=f"{title}!A1",
        valueInputOption="USER_ENTERED",
        body={"values": rows},
    ).execute()


def format_header_and_freeze(svc, sheet_id_num: int, col_count: int):
    requests = [
        {
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id_num, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id_num,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": col_count,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.10, "green": 0.07, "blue": 0.03},
                        "textFormat": {"foregroundColor": {"red": 0.97, "green": 0.84, "blue": 0.45}, "bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat)",
            }
        },
    ]
    svc.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body={"requests": requests}).execute()


def configure_inventory_sheet(svc, sheet_id_num: int):
    requests = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id_num,
                    "startRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 10,
                },
                "cell": {"dataValidation": None},
                "fields": "dataValidation",
            }
        },
        {
            "setDataValidation": {
                "range": {
                    "sheetId": sheet_id_num,
                    "startRowIndex": 1,
                    "startColumnIndex": 4,
                    "endColumnIndex": 5,
                },
                "rule": {
                    "condition": {
                        "type": "ONE_OF_LIST",
                        "values": [
                            {"userEnteredValue": "Disponible"},
                            {"userEnteredValue": "Reservado"},
                            {"userEnteredValue": "Vendido"},
                        ],
                    },
                    "showCustomUi": True,
                    "strict": True,
                },
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id_num,
                    "startRowIndex": 1,
                    "startColumnIndex": 3,
                    "endColumnIndex": 4,
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "CURRENCY",
                            "pattern": "$#,##0.00",
                        }
                    }
                },
                "fields": "userEnteredFormat.numberFormat",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id_num,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": 10,
                },
                "properties": {
                    "pixelSize": 110,
                },
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id_num, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 2},
                "properties": {"pixelSize": 320},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id_num, "dimension": "COLUMNS", "startIndex": 2, "endIndex": 3},
                "properties": {"pixelSize": 140},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id_num, "dimension": "COLUMNS", "startIndex": 5, "endIndex": 6},
                "properties": {"pixelSize": 190},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id_num, "dimension": "COLUMNS", "startIndex": 8, "endIndex": 9},
                "properties": {"pixelSize": 150},
                "fields": "pixelSize",
            }
        },
        {
            "updateDimensionProperties": {
                "range": {"sheetId": sheet_id_num, "dimension": "COLUMNS", "startIndex": 9, "endIndex": 10},
                "properties": {"pixelSize": 260},
                "fields": "pixelSize",
            }
        },
    ]
    svc.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body={"requests": requests}).execute()


def configure_simple_sheet(svc, sheet_id_num: int, col_count: int, currency_cols=None, wide_cols=None):
    currency_cols = currency_cols or []
    wide_cols = wide_cols or {}
    requests = [
        {
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id_num,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": col_count,
                },
                "properties": {"pixelSize": 120},
                "fields": "pixelSize",
            }
        }
    ]
    for col_idx in currency_cols:
        requests.append(
            {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id_num,
                        "startRowIndex": 1,
                        "startColumnIndex": col_idx,
                        "endColumnIndex": col_idx + 1,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "numberFormat": {
                                "type": "CURRENCY",
                                "pattern": "$#,##0.00",
                            }
                        }
                    },
                    "fields": "userEnteredFormat.numberFormat",
                }
            }
        )
    for col_idx, pixel_size in wide_cols.items():
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id_num,
                        "dimension": "COLUMNS",
                        "startIndex": col_idx,
                        "endIndex": col_idx + 1,
                    },
                    "properties": {"pixelSize": pixel_size},
                    "fields": "pixelSize",
                }
            }
        )
    svc.spreadsheets().batchUpdate(spreadsheetId=SHEET_ID, body={"requests": requests}).execute()


def set_sheet_visibility(svc, title: str, hidden: bool):
    meta = svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    for sheet in meta["sheets"]:
        props = sheet["properties"]
        if props["title"] != title:
            continue
        current_hidden = props.get("hidden", False)
        if current_hidden == hidden:
            return
        svc.spreadsheets().batchUpdate(
            spreadsheetId=SHEET_ID,
            body={
                "requests": [
                    {
                        "updateSheetProperties": {
                            "properties": {"sheetId": props["sheetId"], "hidden": hidden},
                            "fields": "hidden",
                        }
                    }
                ]
            },
        ).execute()
        return


def main():
    inventory_rows, stats, missing_photos = build_inventory_rows()
    sales_rows = build_sales_rows()
    reservation_rows = build_reservation_rows(inventory_rows)
    summary_rows = build_summary_rows(inventory_rows, sales_rows)

    inventory_values = [[
        "Artículo",
        "Nombre",
        "Categoría",
        "Precio USD",
        "Estado",
        "Reservado / Comprador",
        "Ref Sotheby's",
        "Página",
        "Estimación Sotheby's",
        "Notas",
    ]] + inventory_rows

    reservation_values = [[
        "Artículo",
        "Nombre",
        "Categoría",
        "Reservado Para",
        "Precio USD",
        "Notas",
    ]] + reservation_rows

    sales_values = [[
        "Artículo",
        "Nombre",
        "Categoría",
        "Comprador",
        "Precio USD",
        "Precio CRC",
        "Fecha venta",
        "Notas",
    ]] + sales_rows

    svc = get_service()

    inv_id = ensure_sheet(svc, "INVENTARIO_MAESTRO")
    res_id = ensure_sheet(svc, "RESERVAS")
    ven_id = ensure_sheet(svc, "VENTAS")
    sum_id = ensure_sheet(svc, "RESUMEN")

    write_sheet(svc, "INVENTARIO_MAESTRO", inventory_values)
    write_sheet(svc, "RESERVAS", reservation_values)
    write_sheet(svc, "VENTAS", sales_values)
    write_sheet(svc, "RESUMEN", summary_rows)

    format_header_and_freeze(svc, inv_id, len(inventory_values[0]))
    format_header_and_freeze(svc, res_id, len(reservation_values[0]))
    format_header_and_freeze(svc, ven_id, len(sales_values[0]))
    format_header_and_freeze(svc, sum_id, len(summary_rows[0]))
    configure_inventory_sheet(svc, inv_id)
    configure_simple_sheet(
        svc,
        res_id,
        len(reservation_values[0]),
        currency_cols=[4],
        wide_cols={1: 320, 3: 180, 5: 260},
    )
    configure_simple_sheet(
        svc,
        ven_id,
        len(sales_values[0]),
        currency_cols=[4],
        wide_cols={1: 320, 3: 180, 5: 140, 7: 260},
    )
    configure_simple_sheet(
        svc,
        sum_id,
        len(summary_rows[0]),
        wide_cols={0: 220, 1: 140},
    )

    for title in ["INVENTARIO_MAESTRO", "RESERVAS", "VENTAS", "RESUMEN"]:
        set_sheet_visibility(svc, title, hidden=False)
    for title in ["IM_Resumen", "IM_Inventario Completo", "IM_Ventas", "IM_Reservas Herederos", "IM_Sothebys"]:
        set_sheet_visibility(svc, title, hidden=True)

    log = {
        "sourceWorkbook": str(MASTER_XLSX.name),
        "sourceClientsWorkbook": str(CLIENTS_XLSX.name),
        "excludedCodes": sorted(EXCLUDED_CODES),
        "inventoryCount": len(inventory_rows),
        "stateCounts": dict(stats),
        "reservationCount": len(reservation_rows),
        "salesCount": len(sales_rows),
        "missingPhotos": missing_photos,
    }
    LOG_PATH.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")

    print("Sincronización completada")
    print(json.dumps(log, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
