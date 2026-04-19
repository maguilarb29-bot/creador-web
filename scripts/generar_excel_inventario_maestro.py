"""
Excel Inventario Maestro Solaris — Sucesión Pignatelli.
518 artículos del catálogo completo con estado, Sotheby's y reservas.
Misma estética que Reservas Herederos y Sotheby's Completo.
"""

import json, re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE     = Path(__file__).resolve().parent.parent
OUT_PATH = BASE / "docs" / "Inventario_Maestro_Solaris_Pignatelli_2026-04-14.xlsx"

# ── Paleta idéntica a los otros Excel ─────────────────────────────────────────
NAVY     = "1A1A2E"
SLATE    = "2C3E50"
GRAY_HDR = "37474F"
GOLD     = "B8860B"
GOLD_LITE= "FEF9E7"
WHITE    = "FFFFFF"

CAT_COLORS = {
    "Arte en papel":    {"dark": "5D4037", "light": "FFF8E1"},
    "Ceramica":         {"dark": "1565C0", "light": "E3F2FD"},
    "Cristaleria":      {"dark": "00695C", "light": "E0F2F1"},
    "Decorativos":      {"dark": "4A148C", "light": "F3E5F5"},
    "Electrodomesticos":{"dark": "424242", "light": "F5F5F5"},
    "Insumos Medicos":  {"dark": "558B2F", "light": "F1F8E9"},
    "Joyas":            {"dark": "880E4F", "light": "FCE4EC"},
    "Muebles":          {"dark": "2E7D32", "light": "E8F5E9"},
    "Plateria":         {"dark": "37474F", "light": "ECEFF1"},
    "Utensilios":       {"dark": "E65100", "light": "FFF3E0"},
}

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, size=10, bold=True, center=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=bold, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border    = thin_border()

def dat(ws, row, col, val, bg=WHITE, fg="1A1A1A", bold=False,
        italic=False, center=False, size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=bold, italic=italic, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border    = thin_border()

def merge_title(ws, row, cols, val, bg=NAVY, fg=WHITE, size=14, height=38):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=val)
    c.font      = Font(name="Calibri", size=size, bold=True, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height

def merge_sub(ws, row, cols, val, bg=SLATE, fg=WHITE, size=10, height=20, italic=True):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=val)
    c.font      = Font(name="Calibri", size=size, bold=False, italic=italic, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height

def fmt_est(s):
    if not s or s in ["-", ""]:
        return "—"
    return s

# ── Cargar catálogo ────────────────────────────────────────────────────────────
with open(BASE / "Api_PG/data/solaris_catalogo.json", encoding="utf-8-sig") as f:
    catalogo = json.load(f)

# Solo ARTICULO y SET (excluir FOTO y LOTE_FOTO que son hijos)
articulos = [a for a in catalogo
             if a.get("tipoEstructural") in ("ARTICULO", "SET", "LOTE")]
articulos.sort(key=lambda x: (x.get("categoria", ""), x.get("codigoItem", "")))

by_cat = defaultdict(list)
for a in articulos:
    by_cat[a["categoria"]].append(a)

# ── Lookup Sotheby's desde CSV (fuente real de estimaciones) ──────────────────
import csv as _csv_mod
_soth_lookup = {}
with open(BASE / "docs/sothebys_maestro_2026-04-13.csv", newline="", encoding="utf-8-sig") as f:
    for row in _csv_mod.DictReader(f):
        est = row.get("estimacion","").strip()
        for parte in row.get("codigo_actual","").split(";"):
            cod = re.sub(r"-\d+$", "", parte.strip()).strip()
            if cod and est and est != "-":
                _soth_lookup[cod] = est

def get_soth_est(codigo):
    return _soth_lookup.get(codigo, "")


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — RESUMEN POR CATEGORÍA
# ══════════════════════════════════════════════════════════════════════════════
def kpi_box(ws, row, col, label, value, bg, fg_val=WHITE, fg_lbl="D0D0D0", colspan=1):
    """Celda KPI con label arriba y valor grande abajo — simulado en 2 filas."""
    # Fila label
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+colspan-1)
    c = ws.cell(row=row, column=col, value=label)
    c.font      = Font(name="Calibri", size=9, bold=False, italic=True, color=fg_lbl)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="bottom")
    c.border    = thin_border()
    # Fila valor
    if colspan > 1:
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+colspan-1)
    c2 = ws.cell(row=row+1, column=col, value=value)
    c2.font      = Font(name="Calibri", size=18, bold=True, color=fg_val)
    c2.fill      = fill(bg)
    c2.alignment = Alignment(horizontal="center", vertical="top")
    c2.border    = thin_border()

def hoja_resumen(wb):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    COLS = 8
    for col, w in zip("ABCDEFGH", [18, 14, 14, 14, 14, 14, 14, 16]):
        ws.column_dimensions[col].width = w

    # ── Encabezado ────────────────────────────────────────────────────────────
    merge_title(ws, 1, COLS,
        "INVENTARIO MAESTRO SOLARIS — SUCESIÓN PIGNATELLI",
        bg=NAVY, size=16, height=44)
    merge_sub(ws, 2, COLS,
        "Condominio Solaris  ·  Inventario completo documentado  ·  Abril 2026",
        bg=SLATE, height=22)
    merge_sub(ws, 3, COLS,
        "Valoraciones Sotheby's International Realty — Londres, 18 de enero de 2016",
        bg="3D5166", fg="D0D8E0", size=9, height=16)

    # ── Calcular totales ──────────────────────────────────────────────────────
    tot_art  = len(articulos)
    vendidos = [a for a in articulos if a.get("estado") == "Vendido"]
    reservados_list = [a for a in articulos if a.get("reservadoPara") and a.get("estado") != "Vendido"]
    n_vend   = len(vendidos)
    n_res    = len(reservados_list)
    n_disp   = tot_art - n_vend - n_res
    n_soth   = sum(1 for a in articulos if a.get("tieneSothebys"))
    tot_usd  = sum(a.get("precioUSD") or 0 for a in vendidos)
    tot_crc  = sum(a.get("precioColones") or 0 for a in vendidos)

    tot_min = tot_max = 0
    for a in articulos:
        est = get_soth_est(a["codigoItem"])
        if est and est not in ["-",""]:
            nums = re.findall(r"\d[\d,]*", est.replace(",",""))
            vals = [int(n) for n in nums if n]
            if vals:
                tot_min += vals[0]
                tot_max += vals[-1]

    # ── Separador ─────────────────────────────────────────────────────────────
    ws.merge_cells("A4:H4")
    ws["A4"].fill = fill("EEF2F7")
    ws.row_dimensions[4].height = 10

    # ── KPIs — fila de labels (5) + valores (6) ───────────────────────────────
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 38

    kpi_box(ws, 5, 1, "TOTAL ARTÍCULOS",  str(tot_art),  "1A1A2E", colspan=2)
    kpi_box(ws, 5, 3, "DISPONIBLES",      str(n_disp),   "1E8449", colspan=2)
    kpi_box(ws, 5, 5, "RESERVADOS",       str(n_res),    "C0392B", colspan=2)
    kpi_box(ws, 5, 7, "VENDIDOS",         str(n_vend),   "7B1FA2", colspan=2)

    # ── Separador ─────────────────────────────────────────────────────────────
    ws.merge_cells("A7:H7")
    ws["A7"].fill = fill("EEF2F7")
    ws.row_dimensions[7].height = 10

    # ── Franja ventas ─────────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 28

    fechas_venta = sorted({a.get("fechaVenta","") for a in vendidos if a.get("fechaVenta","")})
    fecha_rango  = f"{fechas_venta[0]} — {fechas_venta[-1]}" if len(fechas_venta) > 1 else (fechas_venta[0] if fechas_venta else "")
    kpi_box(ws, 8, 1, "VENTAS REALIZADAS", f"{n_vend} artículos · {fecha_rango}",
            "6A1B9A", fg_val="FFFFFF", fg_lbl="D0B0E0", colspan=2)
    kpi_box(ws, 8, 3, "TOTAL EN USD",  f"$ {tot_usd:,.2f}",  "7B1FA2",
            fg_val="FFFFFF", fg_lbl="D0B0E0", colspan=3)
    kpi_box(ws, 8, 6, "TOTAL EN CRC",  f"CRC {tot_crc:,.2f}", "7B1FA2",
            fg_val="FFFFFF", fg_lbl="D0B0E0", colspan=3)

    # ── Franja Sotheby's ──────────────────────────────────────────────────────
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 28

    kpi_box(ws, 10, 1, "ARTÍCULOS CON TASACIÓN SOTHEBY'S", f"{n_soth} artículos",
            "0D47A1", fg_val="FFFFFF", fg_lbl="BBDEFB", colspan=2)
    kpi_box(ws, 10, 3, "ESTIMACIÓN MÍNIMA",  f"$ {tot_min:,}",  "1565C0",
            fg_val="FFFFFF", fg_lbl="BBDEFB", colspan=3)
    kpi_box(ws, 10, 6, "ESTIMACIÓN MÁXIMA",  f"$ {tot_max:,}",  "1565C0",
            fg_val="FFFFFF", fg_lbl="BBDEFB", colspan=3)

    # ── Separador ─────────────────────────────────────────────────────────────
    ws.merge_cells("A12:H12")
    ws["A12"].fill = fill("EEF2F7")
    ws.row_dimensions[12].height = 14

    # ── Tabla por categoría ───────────────────────────────────────────────────
    for col, h in enumerate(
        ["Categoría", "Artículos", "Disponibles", "Reservados",
         "Vendidos", "Con Sotheby's", "Est. Mín. USD", "Est. Máx. USD"], 1):
        hdr(ws, 13, col, h, bg=GRAY_HDR, size=9)
    ws.row_dimensions[13].height = 28

    row_n = 14
    for cat in sorted(by_cat.keys()):
        items  = by_cat[cat]
        colors = CAT_COLORS.get(cat, {"dark": "444444", "light": "F5F5F5"})
        bg     = colors["light"]

        n_v  = sum(1 for a in items if a.get("estado") == "Vendido")
        n_r  = sum(1 for a in items if a.get("reservadoPara") and a.get("estado") != "Vendido")
        n_d  = len(items) - n_v - n_r
        n_s  = sum(1 for a in items if a.get("tieneSothebys"))
        e_min = e_max = 0
        for a in items:
            est = get_soth_est(a["codigoItem"])
            if est and est not in ["-",""]:
                nums = re.findall(r"\d[\d,]*", est.replace(",",""))
                vals = [int(n) for n in nums if n]
                if vals:
                    e_min += vals[0]
                    e_max += vals[-1]

        dat(ws, row_n, 1, cat,         bg=bg, bold=True, fg=colors["dark"])
        dat(ws, row_n, 2, len(items),  bg=bg, center=True, bold=True)
        dat(ws, row_n, 3, n_d if n_d else "—",  bg=bg, center=True,
            fg="1E8449" if n_d else "AAAAAA", bold=bool(n_d))
        dat(ws, row_n, 4, n_r if n_r else "—",  bg=bg, center=True,
            fg="C0392B" if n_r else "AAAAAA", bold=bool(n_r))
        dat(ws, row_n, 5, n_v if n_v else "—",  bg=bg, center=True,
            fg="7B1FA2" if n_v else "AAAAAA", bold=bool(n_v))
        dat(ws, row_n, 6, n_s if n_s else "—",  bg=bg, center=True,
            fg=GOLD if n_s else "AAAAAA", bold=bool(n_s))
        dat(ws, row_n, 7, f"$ {e_min:,}" if e_min else "—",
            bg=bg, center=True, fg="1F4E79", bold=bool(e_min))
        dat(ws, row_n, 8, f"$ {e_max:,}" if e_max else "—",
            bg=bg, center=True, fg="1F4E79", bold=bool(e_max))
        ws.row_dimensions[row_n].height = 22
        row_n += 1

    # Total tabla
    for col, val in enumerate([
        "TOTAL", f"{tot_art}", f"{n_disp}", f"{n_res}",
        f"{n_vend}", f"{n_soth}", f"$ {tot_min:,}", f"$ {tot_max:,}"
    ], 1):
        hdr(ws, row_n, col, val, bg=SLATE, size=10)
    ws.row_dimensions[row_n].height = 22
    row_n += 2

    # Nota al pie
    ws.merge_cells(f"A{row_n}:H{row_n}")
    c = ws[f"A{row_n}"]
    c.value = (
        f"Inventario documentado: {tot_art} artículos — Condominio Solaris, Sucesión Pignatelli.  "
        f"Valoraciones Sotheby's International Realty, Londres, enero 2016. Carácter exclusivamente referencial.  "
        f"Ventas registradas a abril 2026."
    )
    c.font      = Font(name="Calibri", size=9, italic=True, color="777777")
    c.fill      = fill(GOLD_LITE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = thin_border()
    ws.row_dimensions[row_n].height = 36


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — INVENTARIO COMPLETO
# N° | Código | Descripción | Categoría | Ref. S. | Estimación | Estado | Reservado | Notas
# ══════════════════════════════════════════════════════════════════════════════
def hoja_inventario(wb):
    ws = wb.create_sheet("Inventario Completo")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    COLS = 10
    for col, w in zip("ABCDEFGHIJ", [5, 10, 42, 12, 14, 18, 18, 14, 22, 26]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, COLS,
        "INVENTARIO COMPLETO — CONDOMINIO SOLARIS · Sucesión Pignatelli",
        bg=NAVY, size=14, height=32)
    merge_sub(ws, 2, COLS,
        "Todos los artículos documentados  ·  Ordenados por categoría  ·  Abril 2026",
        bg=SLATE, height=20)
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(
        ["N°", "CÓDIGO", "DESCRIPCIÓN DEL OBJETO", "CATEGORÍA",
         "VALORACIÓN\nSOTHEBY'S", "PRECIO\nSUGERIDO (USD)",
         "ESTADO", "RESERVADO / COMPRADOR", "NOTAS"], 1):
        hdr(ws, 3, col, h, bg=GRAY_HDR, size=9)
    # Columna J — header especial para precio sugerido con nota
    c_ph = ws.cell(row=3, column=10, value="PRECIO\nSUGERIDO (USD)")
    c_ph.font      = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c_ph.fill      = fill("455A64")   # gris más claro — indica "pendiente"
    c_ph.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_ph.border    = thin_border()
    ws.row_dimensions[3].height = 28
    ws.auto_filter.ref = "A3:J3"

    row_n = 4
    num   = 0

    for cat in sorted(by_cat.keys()):
        items = by_cat[cat]
        colors = CAT_COLORS.get(cat, {"dark": "444444", "light": "F5F5F5"})

        # Banda de categoría
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
        c = ws[f"A{row_n}"]
        c.value     = f"  {cat.upper()}  -  {len(items)} articulo{'s' if len(items)>1 else ''}"
        c.font      = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill      = fill(colors["dark"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_n].height = 22
        row_n += 1

        for idx, art in enumerate(items):
            num += 1
            bg         = colors["light"] if idx % 2 == 0 else WHITE
            cod        = art["codigoItem"]
            soth_est   = get_soth_est(cod)              # desde CSV — fuente real
            tiene_soth = bool(soth_est)
            precio_sug = art.get("precioUSD")           # precio sugerido Alejandro
            persona    = art.get("reservadoPara", "")
            estado     = art.get("estado", "Disponible")

            if estado == "Vendido":
                est_color, est_txt = "7B1FA2", "Vendido"
            elif persona:
                est_color, est_txt = "C0392B", "Reservado"
            else:
                est_color, est_txt = "27AE60", "Disponible"

            soth_txt  = soth_est if soth_est else "—"
            prec_txt  = f"$ {precio_sug:,.0f}" if precio_sug and estado != "Vendido" else "—"

            dat(ws, row_n, 1,  num,                                    bg=bg, center=True, fg="888888")
            dat(ws, row_n, 2,  cod,                                    bg=bg, center=True, bold=True, fg=colors["dark"])
            dat(ws, row_n, 3,  art.get("nombreES","").capitalize()[:60], bg=bg)
            dat(ws, row_n, 4,  cat,                                    bg=bg, fg="666666", size=9)
            dat(ws, row_n, 5,  soth_txt,                               bg=bg, center=True,
                bold=tiene_soth, fg=GOLD if tiene_soth else "CCCCCC", italic=not tiene_soth)
            dat(ws, row_n, 6,  prec_txt,                               bg=bg, center=True,
                bold=bool(precio_sug and estado != "Vendido"),
                fg="1F4E79" if prec_txt != "—" else "CCCCCC", italic=(prec_txt == "—"))
            dat(ws, row_n, 7,  est_txt,                                bg=bg, center=True, bold=True, fg=est_color)
            dat(ws, row_n, 8,  persona if persona else "—",            bg=bg,
                bold=bool(persona), fg="1A1A2E" if persona else "AAAAAA", italic=not bool(persona))
            dat(ws, row_n, 9,  art.get("notas","") or "",              bg=bg, fg="777777", italic=True, size=9)
            ws.row_dimensions[row_n].height = 20
            row_n += 1

    # Total final
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
    c = ws[f"A{row_n}"]
    c.value     = f"  TOTAL - {num} articulos en el Inventario Maestro Solaris"
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border()
    ws.row_dimensions[row_n].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — VENTAS REALIZADAS
# ══════════════════════════════════════════════════════════════════════════════
def hoja_ventas(wb):
    ws = wb.create_sheet("Ventas")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    COLS = 9
    for col, w in zip("ABCDEFGHI", [5, 10, 40, 22, 14, 14, 18, 14, 36]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, COLS,
        "VENTAS REALIZADAS — CONDOMINIO SOLARIS · Sucesión Pignatelli",
        bg="7B1FA2", size=14, height=32)
    merge_sub(ws, 2, COLS,
        "Artículos del inventario vendidos  ·  Precios en USD y CRC  ·  Actualizado abril 2026",
        bg="6A1B9A", height=20)
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(
        ["N°", "CÓDIGO", "DESCRIPCIÓN DEL OBJETO",
         "COMPRADOR", "CATEGORÍA", "PRECIO (USD)", "PRECIO (CRC)", "FECHA VENTA", "FOTO PRINCIPAL"], 1):
        hdr(ws, 3, col, h, bg=GRAY_HDR, size=9)
    ws.row_dimensions[3].height = 26

    vendidos = [a for a in articulos if a.get("estado") == "Vendido"]
    vendidos.sort(key=lambda x: x.get("codigoItem", ""))

    row_n = 4
    total_usd = 0.0
    total_crc = 0.0
    for idx, art in enumerate(vendidos):
        bg    = "F3E5F5" if idx % 2 == 0 else WHITE
        p_usd = art.get("precioUSD") or 0
        p_crc = art.get("precioColones") or 0
        fecha = art.get("fechaVenta") or "—"
        fotos = art.get("fotos", [])
        foto  = fotos[0] if fotos else "—"
        cat   = art.get("categoria", "")
        total_usd += p_usd
        total_crc += p_crc

        dat(ws, row_n, 1, idx + 1,                                  bg=bg, center=True, fg="888888")
        dat(ws, row_n, 2, art["codigoItem"],                        bg=bg, center=True, bold=True, fg="7B1FA2")
        dat(ws, row_n, 3, art.get("nombreES","").capitalize()[:60], bg=bg)
        dat(ws, row_n, 4, art.get("reservadoPara",""),              bg=bg, bold=True, fg="1A1A2E")
        dat(ws, row_n, 5, cat,                                      bg=bg, fg="666666", size=9)
        dat(ws, row_n, 6, f"$ {p_usd:,.2f}" if p_usd else "—",    bg=bg, center=True, bold=True, fg="1F4E79")
        dat(ws, row_n, 7, f"CRC {p_crc:,.2f}" if p_crc else "—",  bg=bg, center=True, bold=True, fg="27AE60")
        dat(ws, row_n, 8, fecha,                                    bg=bg, center=True,
            fg="555555", italic=(fecha == "—"))
        dat(ws, row_n, 9, foto,                                     bg=bg, fg="AAAAAA", italic=True, size=8)
        ws.row_dimensions[row_n].height = 20
        row_n += 1

    # Total
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=5)
    c = ws[f"A{row_n}"]
    c.value     = f"  TOTAL VENTAS - {len(vendidos)} articulos vendidos"
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill("7B1FA2")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border()
    for col, val in [(6, f"$ {total_usd:,.2f}"), (7, f"CRC {total_crc:,.2f}"), (8, ""), (9, "")]:
        cx = ws.cell(row=row_n, column=col, value=val)
        cx.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cx.fill      = fill("7B1FA2")
        cx.alignment = Alignment(horizontal="center", vertical="center")
        cx.border    = thin_border()
    ws.row_dimensions[row_n].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 4 — ARTÍCULOS CON VALORACIÓN SOTHEBY'S (resumen cruzado)
# ══════════════════════════════════════════════════════════════════════════════
def hoja_sothebys(wb):
    ws = wb.create_sheet("Con Valoración Sotheby's")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    COLS = 7
    for col, w in zip("ABCDEFG", [5, 10, 44, 14, 6, 20, 24]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, COLS,
        "ARTÍCULOS CON VALORACIÓN SOTHEBY'S — Sucesión Pignatelli",
        bg=NAVY, size=14, height=32)
    merge_sub(ws, 2, COLS,
        "Sotheby's International Realty · Londres · 18 de enero de 2016  |  Valores en USD",
        bg=SLATE, height=20)
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(
        ["N°", "CÓDIGO", "DESCRIPCIÓN DEL OBJETO",
         "REF. SOTHEBY'S", "PÁG.", "ESTIMACIÓN (USD)", "ESTADO / RESERVADO"], 1):
        hdr(ws, 3, col, h, bg=GRAY_HDR, size=9)
    ws.row_dimensions[3].height = 26
    ws.auto_filter.ref = "A3:G3"

    con_soth = [a for a in articulos if a.get("tieneSothebys")]
    con_soth.sort(key=lambda x: (x.get("categoria",""), x.get("codigoItem","")))

    row_n = 4
    num = 0
    cat_actual = None

    for art in con_soth:
        cat = art.get("categoria","")
        colors = CAT_COLORS.get(cat, {"dark": "444444", "light": "F5F5F5"})

        if cat != cat_actual:
            cat_count = sum(1 for a in con_soth if a.get("categoria") == cat)
            ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
            c = ws[f"A{row_n}"]
            c.value     = f"  {cat.upper()}  -  {cat_count} articulo{'s' if cat_count>1 else ''}"
            c.font      = Font(name="Calibri", size=11, bold=True, color=WHITE)
            c.fill      = fill(colors["dark"])
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_n].height = 22
            row_n += 1
            cat_actual = cat

        num += 1
        idx_cat = sum(1 for a in con_soth[:con_soth.index(art)] if a.get("categoria") == cat)
        bg  = colors["light"] if idx_cat % 2 == 0 else WHITE
        cod = art["codigoItem"]
        est = get_soth_est(cod)          # desde CSV — fuente real
        ref = art.get("refSothebys","")
        pag = art.get("paginaSothebys","")
        res = art.get("reservadoPara","")
        est_estado = art.get("estado","Disponible")
        if est_estado == "Vendido":
            estado_txt, est_fg = "Vendido",    "7B1FA2"
        elif res:
            estado_txt, est_fg = res,           "C0392B"
        else:
            estado_txt, est_fg = "Disponible", "27AE60"

        dat(ws, row_n, 1, num,                              bg=bg, center=True, fg="888888")
        dat(ws, row_n, 2, cod,                              bg=bg, center=True, bold=True,
            fg=colors["dark"])
        dat(ws, row_n, 3, art.get("nombreES","").capitalize()[:60], bg=bg)
        dat(ws, row_n, 4, f"Ref. {ref}" if ref else "",    bg=bg, center=True,
            bold=bool(ref), fg=GOLD if ref else "AAAAAA")
        dat(ws, row_n, 5, str(pag) if pag else "",         bg=bg, center=True, fg="777777")
        dat(ws, row_n, 6, est if est else "s/d",           bg=bg, center=True,
            bold=bool(est), fg="1F4E79" if est else "AAAAAA",
            italic=not bool(est))
        dat(ws, row_n, 7, estado_txt,                      bg=bg,
            bold=(est_estado == "Vendido" or bool(res)),
            fg=est_fg, italic=(estado_txt == "Disponible"))
        ws.row_dimensions[row_n].height = 20
        row_n += 1

    # Total
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
    c = ws[f"A{row_n}"]
    c.value     = f"  TOTAL - {num} articulos con valoracion Sotheby's"
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border()
    ws.row_dimensions[row_n].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 5 — RESERVAS DE HEREDEROS
# ══════════════════════════════════════════════════════════════════════════════
import csv as _csv

HEREDERO_COLORS = {
    "Adriano Pignatelli":   {"dark": "1A5276", "light": "D6EAF8"},
    "Diego Pignatelli":     {"dark": "1E8449", "light": "D5F5E3"},
    "Fabrizia Pignatelli":  {"dark": "7D3C98", "light": "E8DAEF"},
    "Margherita Pignatelli":{"dark": "B7770D", "light": "FDEBD0"},
    "Maria Cristina Smith": {"dark": "922B21", "light": "FADBD8"},
}

def hoja_reservas(wb):
    ws = wb.create_sheet("Reservas de Herederos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    COLS = 6
    for col, w in zip("ABCDEF", [5, 10, 44, 26, 20, 20]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, COLS,
        "RESERVAS DE HEREDEROS — Sucesión Pignatelli",
        bg=NAVY, size=14, height=32)
    merge_sub(ws, 2, COLS,
        "Artículos reservados por cada heredero  ·  Valoraciones Sotheby's enero 2016",
        bg=SLATE, height=20)
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(
        ["N°", "CÓDIGO", "DESCRIPCIÓN DEL OBJETO",
         "HEREDERO", "CATEGORÍA", "ESTIMACIÓN (USD)"], 1):
        hdr(ws, 3, col, h, bg=GRAY_HDR, size=9)
    ws.row_dimensions[3].height = 26
    ws.auto_filter.ref = "A3:F3"

    # Leer CSV de reservas
    reservas_path = BASE / "docs/reservas_herederos_maestro_2026-04-13.csv"
    soth_path     = BASE / "docs/sothebys_maestro_2026-04-13.csv"

    with open(reservas_path, newline="", encoding="utf-8-sig") as f:
        reservas = list(_csv.DictReader(f))
    with open(soth_path, newline="", encoding="utf-8-sig") as f:
        soth_rows = list(_csv.DictReader(f))

    soth_by_cod = {}
    for r in soth_rows:
        for parte in r.get("codigo_actual","").split(";"):
            cod = re.sub(r"-\d+$", "", parte.strip()).strip()
            if cod:
                soth_by_cod[cod] = r.get("estimacion","").strip()

    # Agrupar por heredero
    from collections import defaultdict as _dd
    por_heredero = _dd(list)
    for r in reservas:
        por_heredero[r["heredero"]].append(r)

    row_n = 4
    num   = 0
    cat_lk = {a["codigoItem"]: a.get("categoria","") for a in articulos}

    for heredero in sorted(por_heredero.keys()):
        items_h = por_heredero[heredero]
        colors  = HEREDERO_COLORS.get(heredero, {"dark": "444444", "light": "F5F5F5"})

        # Banda de heredero
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
        c = ws[f"A{row_n}"]
        c.value     = f"  {heredero.upper()}  -  {len(items_h)} articulo{'s' if len(items_h)>1 else ''}"
        c.font      = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill      = fill(colors["dark"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_n].height = 22
        row_n += 1

        for idx, r in enumerate(items_h):
            num += 1
            cod  = r["codigo"].strip()
            bg   = colors["light"] if idx % 2 == 0 else WHITE
            est  = soth_by_cod.get(cod, "—")
            if est in ["-", ""]: est = "—"
            cat  = cat_lk.get(cod, "")

            dat(ws, row_n, 1, num,                     bg=bg, center=True, fg="888888")
            dat(ws, row_n, 2, cod,                     bg=bg, center=True, bold=True, fg=colors["dark"])
            dat(ws, row_n, 3, r["nombre"].capitalize()[:60], bg=bg)
            dat(ws, row_n, 4, heredero,                bg=bg, bold=True, fg=colors["dark"])
            dat(ws, row_n, 5, cat,                     bg=bg, fg="666666", size=9)
            dat(ws, row_n, 6, est,                     bg=bg, center=True,
                bold=(est != "—"), fg="1F4E79" if est != "—" else "AAAAAA",
                italic=(est == "—"))
            ws.row_dimensions[row_n].height = 20
            row_n += 1

    # Total
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
    c = ws[f"A{row_n}"]
    c.value     = f"  TOTAL - {num} articulos reservados por herederos"
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border()
    ws.row_dimensions[row_n].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 6 — RESERVAS Y VENTAS DE CLIENTES EXTERNOS
# ══════════════════════════════════════════════════════════════════════════════
CLIENTE_COLORS = {
    "Rasmi":           {"dark": "0D47A1", "light": "E3F2FD"},
    "Nuria":           {"dark": "880E4F", "light": "FCE4EC"},
    "Alejandro":       {"dark": "E65100", "light": "FFF3E0"},
    "Pablo Brenes":    {"dark": "2E7D32", "light": "E8F5E9"},
    "Marielos Alpizar":{"dark": "4A148C", "light": "F3E5F5"},
    "Ana Martin":      {"dark": "00695C", "light": "E0F2F1"},
    "Luciana Paris":   {"dark": "F57F17", "light": "FFFDE7"},
    "Evelina Vargas":  {"dark": "00838F", "light": "E0F7FA"},
}

HEREDEROS = {
    "Adriano Pignatelli","Diego Pignatelli","Fabrizia Pignatelli",
    "Margherita Pignatelli","Maria Cristina Smith",
}

def hoja_clientes(wb):
    ws = wb.create_sheet("Reservas y Ventas Clientes")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    COLS = 8
    for col, w in zip("ABCDEFGH", [5, 10, 44, 22, 14, 14, 18, 26]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, COLS,
        "RESERVAS Y VENTAS CLIENTES EXTERNOS — Sucesión Pignatelli",
        bg="0D47A1", size=14, height=32)
    merge_sub(ws, 2, COLS,
        "Artículos reservados y vendidos a compradores externos  ·  Abril 2026",
        bg="1565C0", height=20)
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(
        ["N°", "CÓDIGO", "DESCRIPCIÓN DEL OBJETO",
         "CLIENTE", "CATEGORÍA", "PRECIO (USD)", "ESTADO", "NOTAS"], 1):
        hdr(ws, 3, col, h, bg=GRAY_HDR, size=9)
    ws.row_dimensions[3].height = 26
    ws.auto_filter.ref = "A3:H3"

    # Agrupar por cliente (excluir herederos)
    from collections import defaultdict as _dd
    por_cliente = _dd(list)
    for a in articulos:
        cli = a.get("reservadoPara","")
        if cli and cli not in HEREDEROS:
            por_cliente[cli].append(a)

    row_n = 4
    num   = 0
    tot_usd = 0.0

    for cli in sorted(por_cliente.keys()):
        items  = por_cliente[cli]
        colors = CLIENTE_COLORS.get(cli, {"dark": "444444", "light": "F5F5F5"})

        # Banda cliente
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
        c = ws[f"A{row_n}"]
        c.value     = f"  {cli.upper()}  -  {len(items)} articulo{'s' if len(items)>1 else ''}"
        c.font      = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill      = fill(colors["dark"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_n].height = 22
        row_n += 1

        for idx, a in enumerate(items):
            num += 1
            bg     = colors["light"] if idx % 2 == 0 else WHITE
            precio = a.get("precioUSD") or 0
            estado = a.get("estado","Disponible")
            if estado == "Vendido":
                est_txt, est_fg = "Vendido",  "7B1FA2"
                tot_usd += precio
            else:
                est_txt, est_fg = "Reservado","C0392B"

            dat(ws, row_n, 1, num,                                       bg=bg, center=True, fg="888888")
            dat(ws, row_n, 2, a["codigoItem"],                           bg=bg, center=True, bold=True, fg=colors["dark"])
            dat(ws, row_n, 3, a.get("nombreES","").capitalize()[:60],    bg=bg, size=9)
            dat(ws, row_n, 4, cli,                                       bg=bg, bold=True, fg=colors["dark"], size=9)
            dat(ws, row_n, 5, a.get("categoria",""),                     bg=bg, fg="666666", size=8, center=True)
            dat(ws, row_n, 6, f"$ {precio:,.0f}" if precio else "—",    bg=bg, center=True,
                bold=bool(precio), fg="1F4E79" if precio else "AAAAAA",
                italic=not bool(precio))
            dat(ws, row_n, 7, est_txt,                                   bg=bg, center=True, bold=True, fg=est_fg)
            dat(ws, row_n, 8, a.get("notas","") or "",                  bg=bg, fg="777777", italic=True, size=8)
            ws.row_dimensions[row_n].height = 20
            row_n += 1

    # Total
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=5)
    c = ws[f"A{row_n}"]
    c.value     = f"  TOTAL - {num} articulos de clientes externos"
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border()
    for col, val in [(6, f"$ {tot_usd:,.2f}" if tot_usd else "—"), (7, ""), (8, "")]:
        cx = ws.cell(row=row_n, column=col, value=val)
        cx.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cx.fill      = fill(SLATE)
        cx.alignment = Alignment(horizontal="center", vertical="center")
        cx.border    = thin_border()
    ws.row_dimensions[row_n].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    hoja_resumen(wb)
    hoja_inventario(wb)
    hoja_ventas(wb)
    hoja_reservas(wb)
    hoja_sothebys(wb)

    hoja_clientes(wb)

    wb.properties.title   = "Inventario Maestro Solaris — Sucesión Pignatelli 2026"
    wb.properties.subject = "Inventario completo del Condominio Solaris"
    wb.properties.creator = "Sistema Pignatelli"
    wb.save(OUT_PATH)

    n_soth    = sum(1 for a in articulos if a.get("tieneSothebys"))
    n_res     = sum(1 for a in articulos if a.get("reservadoPara") and a.get("estado") != "Vendido")
    n_vend    = sum(1 for a in articulos if a.get("estado") == "Vendido")
    tot_ventas= sum(a.get("precioUSD") or 0 for a in articulos if a.get("estado") == "Vendido")

    print(f"\nOK  {OUT_PATH.name}")
    print(f"    Artículos totales : {len(articulos)}")
    print(f"    Con Sotheby's     : {n_soth}")
    print(f"    Reservados        : {n_res}")
    print(f"    Vendidos          : {n_vend}  (USD {tot_ventas:,})")
    print(f"    Hojas             : Resumen | Inventario Completo | Ventas | Reservas Herederos | Con Valoración Sotheby's | Reservas y Ventas Clientes")

if __name__ == "__main__":
    main()
