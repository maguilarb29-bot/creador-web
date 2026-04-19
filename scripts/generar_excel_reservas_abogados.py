"""
Excel profesional de Reservas de Herederos — Sucesión Pignatelli / Condominio Solaris
Diseño limpio para presentación a abogados.
"""

import csv, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR     = Path(__file__).resolve().parent.parent
CSV_RESERVAS = BASE_DIR / "docs" / "reservas_herederos_maestro_2026-04-13.csv"
CSV_SOTHEBYS = BASE_DIR / "docs" / "sothebys_maestro_2026-04-13.csv"
OUT_PATH     = BASE_DIR / "docs" / "Reservas_Herederos_Pignatelli_2026-04-13.xlsx"

# ── Colores ────────────────────────────────────────────────────────────────────
HEREDERO_COLORS = {
    "Adriano Pignatelli":    {"dark": "1F4E79", "light": "DCE9F5"},
    "Diego Pignatelli":      {"dark": "2E6B30", "light": "DFF0DC"},
    "Fabrizia Pignatelli":   {"dark": "7B3F00", "light": "FDEEDE"},
    "Margherita Pignatelli": {"dark": "5B1F5E", "light": "F2E0F4"},
    "Maria Cristina Smith":  {"dark": "1A4480", "light": "D9E8F6"},
}
NAVY      = "1A1A2E"
SLATE     = "2C3E50"
GOLD      = "B8860B"
GOLD_LITE = "FEF9E7"
GRAY_HDR  = "37474F"
GRAY_ROW  = "F7F7F7"
WHITE     = "FFFFFF"

# ── Bordes ─────────────────────────────────────────────────────────────────────
def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color="AAAAAA"):
    return Border(bottom=Side(style="medium", color=color))

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ── Helpers de celda ───────────────────────────────────────────────────────────
def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, size=10, bold=True, center=True, wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=bold, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=wrap)
    c.border    = thin_border()
    return c

def dat(ws, row, col, val, bg=WHITE, fg="1A1A1A", size=10,
        bold=False, italic=False, center=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=bold, italic=italic, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border    = thin_border()
    return c

def merge_title(ws, row, cols, val, bg=NAVY, fg=WHITE, size=14, height=34):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=val)
    c.font      = Font(name="Calibri", size=size, bold=True, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height
    return c

def merge_subtitle(ws, row, cols, val, bg=SLATE, fg=WHITE, size=11, height=22, italic=True):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=val)
    c.font      = Font(name="Calibri", size=size, bold=False, italic=italic, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height
    return c

# ── Datos ──────────────────────────────────────────────────────────────────────
def leer_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))

def parse_estimacion(s):
    nums = re.findall(r"\d[\d,]*", s.replace(",", ""))
    vals = [int(n) for n in nums if n]
    if len(vals) >= 2:
        return vals[0], vals[1]
    if len(vals) == 1:
        return vals[0], vals[0]   # valor único (ej. £ convertidas)
    return None, None

def cargar_sothebys(path):
    lookup = {}
    for row in leer_csv(path):
        for parte in row.get("codigo_actual","").split(";"):
            base = re.sub(r"-\d+$", "", parte.strip()).strip()
            if base:
                lookup[base] = {
                    "ref":   row.get("ref_sothebys","").strip(),
                    "pag":   row.get("pagina","").strip(),
                    "est":   row.get("estimacion","").strip(),
                    "desc":  row.get("descripcion_sothebys","").strip(),
                }
    return lookup

def fmt_est(est):
    """'$7,000 - $9,800' → limpio o vacío."""
    return est if est and est != "-" else ""


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — RESUMEN EJECUTIVO
# ══════════════════════════════════════════════════════════════════════════════
def hoja_resumen(wb, rows, sothebys):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    # Anchos: Heredero | Items | Tasados | Estimado mín–máx | Observación
    for col, w in zip("ABCDEF", [30, 12, 12, 20, 20, 30]):
        ws.column_dimensions[col].width = w

    # Encabezado
    merge_title(ws, 1, 6,
        "SUCESIÓN PIGNATELLI — CONDOMINIO SOLARIS",
        bg=NAVY, size=16, height=42)
    merge_subtitle(ws, 2, 6,
        "Reservas declaradas por heredero sobre el Inventario Solaris  ·  Abril de 2026",
        bg=SLATE, height=26)
    merge_subtitle(ws, 3, 6,
        "Las valoraciones económicas son referenciales y corresponden a la tasación de "
        "Sotheby's International Realty, Londres, 18 de enero de 2016  (valores en USD)",
        bg="3D5166", fg="D0D8E0", size=10, height=28)

    # Espacio
    ws.merge_cells("A4:F4")
    ws["A4"].fill = fill("EEF2F7")
    ws.row_dimensions[4].height = 10

    # Cabecera tabla
    hdrs = ["Heredero", "Objetos\nReservados", "Con Tasación\nSotheby's",
            "Estimado Mínimo\n(USD)", "Estimado Máximo\n(USD)", "Observación"]
    for col, h in enumerate(hdrs, 1):
        hdr(ws, 5, col, h, bg=GRAY_HDR, size=10)
    ws.row_dimensions[5].height = 32

    # Acumular datos
    herederos = {}
    for r in rows:
        h = r["heredero"]
        if h not in herederos:
            herederos[h] = {"items": 0, "tasados": 0, "emin": 0, "emax": 0, "sin": []}
        herederos[h]["items"] += 1
        sv = sothebys.get(r["codigo"], {})
        if sv.get("ref"):
            herederos[h]["tasados"] += 1
            lo, hi = parse_estimacion(sv.get("est",""))
            if lo: herederos[h]["emin"] += lo
            if hi: herederos[h]["emax"] += hi
        else:
            herederos[h]["sin"].append(r["codigo"])

    tot_items = tot_tasados = tot_min = tot_max = 0
    row_n = 6
    for heredero, d in herederos.items():
        colors = HEREDERO_COLORS.get(heredero, {"dark":"555555","light":"EEEEEE"})
        bg = colors["light"]
        emin = f"$ {d['emin']:,}" if d["emin"] else "—"
        emax = f"$ {d['emax']:,}" if d["emax"] else "—"
        sin_n = len(d["sin"])
        obs = (f"{sin_n} objeto{'s' if sin_n>1 else ''} sin tasación Sotheby's"
               if sin_n else "Todos los objetos tienen referencia Sotheby's")

        dat(ws, row_n, 1, heredero,      bg=bg, bold=True, fg=colors["dark"])
        dat(ws, row_n, 2, d["items"],    bg=bg, center=True, bold=True)
        dat(ws, row_n, 3, f"{d['tasados']} de {d['items']}", bg=bg, center=True)
        dat(ws, row_n, 4, emin, bg=bg, center=True, bold=bool(d["emin"]), fg="1F4E79")
        dat(ws, row_n, 5, emax, bg=bg, center=True, bold=bool(d["emax"]), fg="1F4E79")
        dat(ws, row_n, 6, obs,  bg=bg, italic=True, fg="555555")
        ws.row_dimensions[row_n].height = 26

        tot_items += d["items"]; tot_tasados += d["tasados"]
        tot_min   += d["emin"];  tot_max     += d["emax"]
        row_n += 1

    # Fila totales
    for col, val in enumerate([
        "TOTAL GENERAL", f"{tot_items} objetos",
        f"{tot_tasados} tasados",
        f"$ {tot_min:,}", f"$ {tot_max:,}", ""
    ], 1):
        hdr(ws, row_n, col, val, bg=SLATE, size=10)
    ws.row_dimensions[row_n].height = 24
    row_n += 2

    # Nota legal — texto más natural
    ws.merge_cells(f"A{row_n}:F{row_n}")
    c = ws[f"A{row_n}"]
    c.value = (
        "Nota:  Las reservas indicadas en este documento reflejan la intención declarada de cada heredero "
        "sobre los bienes del acervo sucesoral del Condominio Solaris. No constituyen asignación definitiva "
        "de propiedad y están sujetas al proceso formal de partición ante las instancias legales competentes. "
        "Los valores Sotheby's son de enero de 2016 y tienen carácter exclusivamente referencial."
    )
    c.font      = Font(name="Calibri", size=9, italic=True, color="555555")
    c.fill      = fill(GOLD_LITE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = thin_border("D4AC0D")
    ws.row_dimensions[row_n].height = 72


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — DETALLE POR HEREDERO
# Columnas: N° | Código | Descripción | Categoría | Fotos | Ref.Sotheby's | Estimación (USD)
# ══════════════════════════════════════════════════════════════════════════════
def hoja_detalle(wb, rows, sothebys):
    ws = wb.create_sheet("Detalle por Heredero")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"   # congela título + subtítulo

    # A   B       C                D          E      F              G
    # N°  Código  Descripción      Categoría  Fotos  Ref.Sotheby's  Estimación
    for col, w in zip("ABCDEFG", [5, 10, 44, 18, 7, 14, 20]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, 7,
        "DETALLE DE RESERVAS POR HEREDERO — SUCESIÓN PIGNATELLI",
        bg=NAVY, size=14, height=32)
    merge_subtitle(ws, 2, 7,
        "Valoración Sotheby's 2016  |  Inventario fotográfico oficial Condominio Solaris",
        bg=SLATE, height=20)
    ws.row_dimensions[3].height = 6

    row_n = 4
    heredero_items = {}
    for r in rows:
        heredero_items.setdefault(r["heredero"], []).append(r)

    num = 0
    for heredero, items in heredero_items.items():
        colors = HEREDERO_COLORS.get(heredero, {"dark":"444444","light":"EEEEEE"})

        # Banda de heredero — ocupa toda la fila
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=7)
        c = ws[f"A{row_n}"]
        c.value     = f"  {heredero.upper()}"
        c.font      = Font(name="Calibri", size=12, bold=True, color=WHITE)
        c.fill      = fill(colors["dark"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_n].height = 24
        row_n += 1

        # Cabecera de columnas
        COLS_HDR = ["N°", "CÓDIGO", "DESCRIPCIÓN", "CATEGORÍA", "FOTOS",
                    "REF.\nSOTHEBY'S", "ESTIMACIÓN\n(USD)"]
        for col, h in enumerate(COLS_HDR, 1):
            hdr(ws, row_n, col, h, bg=GRAY_HDR, size=9)
        ws.row_dimensions[row_n].height = 26
        row_n += 1

        sub_min = sub_max = 0
        for idx, r in enumerate(items):
            num += 1
            bg  = colors["light"] if idx % 2 == 0 else WHITE
            sv  = sothebys.get(r["codigo"], {})
            ref = sv.get("ref", "") or ""
            est = fmt_est(sv.get("est", ""))
            lo, hi = parse_estimacion(sv.get("est",""))
            if lo: sub_min += lo
            if hi: sub_max += hi

            dat(ws, row_n, 1, num,              bg=bg, center=True, fg="666666")
            dat(ws, row_n, 2, r["codigo"],      bg=bg, center=True, bold=True, fg=colors["dark"])
            dat(ws, row_n, 3, r["nombre"].capitalize(), bg=bg)
            dat(ws, row_n, 4, r["categoria"],   bg=bg, fg="444444")
            dat(ws, row_n, 5, int(r["cantidad_fotos"]), bg=bg, center=True, fg="666666")
            # Ref Sotheby's
            if ref:
                dat(ws, row_n, 6, f"Ref. {ref}", bg=bg, center=True, bold=True, fg=GOLD)
            else:
                dat(ws, row_n, 6, "—", bg=bg, center=True, fg="BBBBBB", italic=True)
            # Estimación
            if est:
                dat(ws, row_n, 7, est, bg=bg, center=True, bold=True, fg="1F4E79")
            else:
                dat(ws, row_n, 7, "—", bg=bg, center=True, fg="BBBBBB", italic=True)

            ws.row_dimensions[row_n].height = 20
            row_n += 1

        # Subtotal heredero
        rango = f"$ {sub_min:,} – $ {sub_max:,}" if sub_min else "Sin tasación disponible"
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=5)
        c = ws[f"A{row_n}"]
        c.value     = f"  {len(items)} ítem{'s' if len(items)>1 else ''} reservado{'s' if len(items)>1 else ''}"
        c.font      = Font(name="Calibri", size=9, italic=True, color="555555")
        c.fill      = fill("F0F4F8")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin_border("CCCCCC")
        ws.merge_cells(start_row=row_n, start_column=6, end_row=row_n, end_column=7)
        c2 = ws.cell(row=row_n, column=6, value=f"Rango: {rango}")
        c2.font      = Font(name="Calibri", size=9, bold=bool(sub_min), color="1F4E79" if sub_min else "888888")
        c2.fill      = fill("F0F4F8")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border    = thin_border("CCCCCC")
        ws.row_dimensions[row_n].height = 18
        row_n += 2   # espacio entre herederos


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — VALORACIÓN SOTHEBY'S  (misma estética que Detalle por Heredero)
# Columnas: N° | Código | Descripción | Categoría | Ref. Sotheby's | Pág. | Estimación
# ══════════════════════════════════════════════════════════════════════════════
def hoja_sothebys(wb, rows, sothebys):
    ws = wb.create_sheet("Valoración Sotheby's")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # A   B       C              D          E              F     G
    # N°  Código  Descripción    Categoría  Ref.Sotheby's  Pág.  Estimación
    for col, w in zip("ABCDEFG", [5, 10, 44, 18, 14, 6, 20]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, 7,
        "VALORACIÓN SOTHEBY'S — ÍTEMS RESERVADOS · Sucesión Pignatelli",
        bg=NAVY, size=14, height=32)
    merge_subtitle(ws, 2, 7,
        "Sotheby's International Realty · Londres · 18 de enero de 2016  |  Valores expresados en USD",
        bg=SLATE, height=20)
    ws.row_dimensions[3].height = 6

    # Cabecera fija
    COLS_HDR = ["N°", "CÓDIGO", "DESCRIPCIÓN DEL OBJETO",
                "CATEGORÍA", "REF.\nSOTHEBY'S", "PÁG.", "ESTIMACIÓN\n(USD)"]
    for col, h in enumerate(COLS_HDR, 1):
        hdr(ws, 3, col, h, bg=GRAY_HDR, size=9)
    ws.row_dimensions[3].height = 26
    ws.auto_filter.ref = "A3:G3"
    ws.freeze_panes = "A4"

    row_n = 4
    g_min = g_max = num = 0

    # Agrupar por heredero — misma estructura que hoja_detalle
    heredero_items = {}
    for r in rows:
        sv  = sothebys.get(r["codigo"], {})
        if not sv.get("ref"):
            continue
        heredero_items.setdefault(r["heredero"], []).append((r, sv))

    for heredero, items in heredero_items.items():
        colors = HEREDERO_COLORS.get(heredero, {"dark":"444444","light":"EEEEEE"})

        # Banda de heredero
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=7)
        c = ws[f"A{row_n}"]
        c.value     = f"  {heredero.upper()}"
        c.font      = Font(name="Calibri", size=12, bold=True, color=WHITE)
        c.fill      = fill(colors["dark"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_n].height = 24
        row_n += 1

        sub_min = sub_max = 0
        for idx, (r, sv) in enumerate(items):
            num += 1
            bg  = colors["light"] if idx % 2 == 0 else WHITE
            ref = sv.get("ref","")
            pag = sv.get("pag","")
            est = fmt_est(sv.get("est",""))
            lo, hi = parse_estimacion(sv.get("est",""))
            if lo: sub_min += lo; g_min += lo
            if hi: sub_max += hi; g_max += hi

            dat(ws, row_n, 1, num,                       bg=bg, center=True, fg="666666")
            dat(ws, row_n, 2, r["codigo"],               bg=bg, center=True, bold=True, fg=colors["dark"])
            dat(ws, row_n, 3, r["nombre"].capitalize(),  bg=bg)
            dat(ws, row_n, 4, r["categoria"],            bg=bg, fg="444444")
            dat(ws, row_n, 5, f"Ref. {ref}",             bg=bg, center=True, bold=True, fg=GOLD)
            dat(ws, row_n, 6, pag,                       bg=bg, center=True, fg="777777")
            if est:
                dat(ws, row_n, 7, est, bg=bg, center=True, bold=True, fg="1F4E79")
            else:
                dat(ws, row_n, 7, "Ref. sin precio", bg=bg, center=True,
                    italic=True, fg="999999")
            ws.row_dimensions[row_n].height = 20
            row_n += 1

        # Subtotal heredero
        rango = f"$ {sub_min:,} – $ {sub_max:,}" if sub_min else "Solo referencia"
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
        c = ws[f"A{row_n}"]
        c.value     = f"  {len(items)} ítem{'s' if len(items)>1 else ''} con referencia Sotheby's"
        c.font      = Font(name="Calibri", size=9, italic=True, color="555555")
        c.fill      = fill("F0F4F8")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin_border("CCCCCC")
        c2 = ws.cell(row=row_n, column=7, value=f"Rango: {rango}")
        c2.font      = Font(name="Calibri", size=9, bold=bool(sub_min), color="1F4E79" if sub_min else "888888")
        c2.fill      = fill("F0F4F8")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border    = thin_border("CCCCCC")
        ws.row_dimensions[row_n].height = 18
        row_n += 2

    # Total general
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
    c = ws[f"A{row_n}"]
    c.value     = f"  TOTAL TASADO  —  {num} ítems con referencia Sotheby's"
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border()
    c2 = ws.cell(row=row_n, column=7, value=f"$ {g_min:,} – $ {g_max:,}")
    c2.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c2.fill      = fill(SLATE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border    = thin_border()
    ws.row_dimensions[row_n].height = 22
    row_n += 2

    # Nota ítems sin tasación
    sin = [r["codigo"] for r in rows if not sothebys.get(r["codigo"],{}).get("ref")]
    if sin:
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=7)
        c = ws[f"A{row_n}"]
        c.value = (
            f"Ítems reservados sin referencia Sotheby's 2016 ({len(sin)}): "
            + ", ".join(sin)
            + "  —  No fueron incluidos en la tasación de enero 2016 "
              "(lotes de libros, joyas de fantasía, cristalería no catalogada)."
        )
        c.font      = Font(name="Calibri", size=9, italic=True, color="444444")
        c.fill      = fill(GOLD_LITE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = thin_border("D4AC0D")
        ws.row_dimensions[row_n].height = 40


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    rows     = leer_csv(CSV_RESERVAS)
    sothebys = cargar_sothebys(CSV_SOTHEBYS)
    wb       = Workbook()

    hoja_resumen(wb, rows, sothebys)
    hoja_detalle(wb, rows, sothebys)
    hoja_sothebys(wb, rows, sothebys)

    wb.properties.title   = "Reservas Herederos Pignatelli 2026"
    wb.properties.subject = "Sucesión Pignatelli — Condominio Solaris"
    wb.properties.creator = "Sistema Pignatelli"

    wb.save(OUT_PATH)

    tasados = [r for r in rows if sothebys.get(r["codigo"],{}).get("ref")]
    g_min   = sum(parse_estimacion(sothebys[r["codigo"]]["est"])[0] or 0 for r in tasados)
    g_max   = sum(parse_estimacion(sothebys[r["codigo"]]["est"])[1] or 0 for r in tasados)

    print(f"\nOK  {OUT_PATH.name}")
    print(f"    Items          : {len(rows)}")
    print(f"    Con tasacion   : {len(tasados)}")
    print(f"    Rango estimado : $  {g_min:,}  –  $  {g_max:,}  USD")
    print(f"    Hojas          : Resumen | Detalle por Heredero | Valoracion Sotheby's")

if __name__ == "__main__":
    main()
