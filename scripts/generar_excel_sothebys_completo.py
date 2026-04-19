"""
Excel Sotheby's completo — misma estética que Reservas Herederos Pignatelli.
57 artículos del inventario Solaris con valoración Sotheby's.
"""

import csv, json, re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE     = Path(__file__).resolve().parent.parent
OUT_PATH = BASE / "docs" / "Sothebys" / "Sothebys_Inventario_Completo_Pignatelli_2026-04-13.xlsx"

# ── Paleta idéntica a Reservas Herederos ──────────────────────────────────────
NAVY     = "1A1A2E"
SLATE    = "2C3E50"
GRAY_HDR = "37474F"
GOLD     = "B8860B"
GOLD_LITE= "FEF9E7"
WHITE    = "FFFFFF"

CAT_COLORS = {
    "Arte en papel":  {"dark": "5D4037", "light": "FFF8E1"},
    "Ceramica":       {"dark": "1565C0", "light": "E3F2FD"},
    "Cristaleria":    {"dark": "00695C", "light": "E0F2F1"},
    "Decorativos":    {"dark": "4A148C", "light": "F3E5F5"},
    "Muebles":        {"dark": "2E7D32", "light": "E8F5E9"},
    "Plateria":       {"dark": "37474F", "light": "ECEFF1"},
    "Joyas":          {"dark": "880E4F", "light": "FCE4EC"},
}

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, size=10, bold=True, center=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=bold, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border    = thin_border()

def dat(ws, row, col, val, bg=WHITE, fg="1A1A1A", bold=False,
        italic=False, center=False, size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=bold, italic=italic, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border    = thin_border()

def merge_title(ws, row, cols, val, bg=NAVY, fg=WHITE, size=14, height=34):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=val)
    c.font      = Font(name="Calibri", size=size, bold=True, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height

def merge_sub(ws, row, cols, val, bg=SLATE, fg=WHITE, size=10, height=20, italic=True):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=val)
    c.font      = Font(name="Calibri", size=size, bold=False, italic=italic, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height

def parse_est(s):
    if not s or s in ["-", ""]:
        return None, None
    nums = re.findall(r"\d[\d,]*", s.replace(",", ""))
    vals = [int(n) for n in nums if n]
    if len(vals) >= 2: return vals[0], vals[1]
    if len(vals) == 1: return vals[0], vals[0]
    return None, None

def leer_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))

# ── Cargar y cruzar datos ──────────────────────────────────────────────────────
with open(BASE / "Api_PG/data/solaris_catalogo.json", encoding="utf-8-sig") as f:
    catalogo = json.load(f)

soth_csv = leer_csv(BASE / "docs/sothebys_maestro_2026-04-13.csv")
reservas = leer_csv(BASE / "docs/reservas_herederos_maestro_2026-04-13.csv")

cat_lookup     = {a["codigoItem"].strip(): a for a in catalogo}
reserva_lookup = {r["codigo"].strip(): r["heredero"] for r in reservas}

def base_cod(raw):
    return re.sub(r"-\d+$", "", raw.split(";")[0].strip()).strip()

# Lista maestra — un item por código
items = []
seen  = set()
for row in soth_csv:
    for parte in row.get("codigo_actual", "").split(";"):
        cod = base_cod(parte)
        if not cod or cod in seen:
            continue
        seen.add(cod)
        art   = cat_lookup.get(cod, {})
        nombre = (art.get("nombreES") or art.get("nombre") or cod).strip()
        cat    = art.get("categoria", "Sin categoría")
        items.append({
            "codigo":    cod,
            "nombre":    nombre[:60],          # máx 60 chars — sin desborde
            "categoria": cat,
            "ref":       row.get("ref_sothebys", "").strip(),
            "pag":       row.get("pagina", "").strip(),
            "est":       row.get("estimacion", "").strip(),
            "reservado": reserva_lookup.get(cod, ""),
        })

items.sort(key=lambda x: (x["categoria"], x["codigo"]))


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — RESUMEN  (igual que Resumen de Reservas)
# ══════════════════════════════════════════════════════════════════════════════
def hoja_resumen(wb):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    for col, w in zip("ABCDEF", [24, 12, 14, 20, 20, 26]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, 6,
        "INVENTARIO SOLARIS — ARTÍCULOS CON VALORACIÓN SOTHEBY'S",
        bg=NAVY, size=15, height=40)
    merge_sub(ws, 2, 6,
        "Sucesión Pignatelli  ·  Sotheby's International Realty, Londres, 18 de enero de 2016",
        bg=SLATE, height=22)
    merge_sub(ws, 3, 6,
        "Documento de referencia económica  ·  Valores en USD",
        bg="3D5166", fg="D0D8E0", size=10, height=18)

    ws.merge_cells("A4:F4")
    ws["A4"].fill = fill("EEF2F7")
    ws.row_dimensions[4].height = 10

    for col, h in enumerate(
        ["Categoría", "Artículos\nTasados", "Con Precio\nEstimado",
         "Estimado Mínimo\n(USD)", "Estimado Máximo\n(USD)", "Observación"], 1):
        hdr(ws, 5, col, h, bg=GRAY_HDR, size=10)
    ws.row_dimensions[5].height = 30

    por_cat = defaultdict(lambda: {"n": 0, "cp": 0, "emin": 0, "emax": 0})
    for item in items:
        cat = item["categoria"]
        por_cat[cat]["n"] += 1
        lo, hi = parse_est(item["est"])
        if lo:
            por_cat[cat]["cp"]   += 1
            por_cat[cat]["emin"] += lo
            por_cat[cat]["emax"] += hi

    tot_n = tot_cp = tot_min = tot_max = 0
    row_n = 6
    for cat, d in sorted(por_cat.items()):
        colors = CAT_COLORS.get(cat, {"dark": "444444", "light": "F5F5F5"})
        bg = colors["light"]
        sin = d["n"] - d["cp"]
        obs = f"{sin} solo referencia, sin precio" if sin else "Todos con precio estimado"

        dat(ws, row_n, 1, cat,  bg=bg, bold=True, fg=colors["dark"])
        dat(ws, row_n, 2, d["n"], bg=bg, center=True, bold=True)
        dat(ws, row_n, 3, f"{d['cp']} de {d['n']}", bg=bg, center=True)
        dat(ws, row_n, 4,
            f"$ {d['emin']:,}" if d["emin"] else "—",
            bg=bg, center=True, bold=bool(d["emin"]), fg="1F4E79")
        dat(ws, row_n, 5,
            f"$ {d['emax']:,}" if d["emax"] else "—",
            bg=bg, center=True, bold=bool(d["emax"]), fg="1F4E79")
        dat(ws, row_n, 6, obs, bg=bg, italic=True, fg="666666", size=9)
        ws.row_dimensions[row_n].height = 24

        tot_n += d["n"]; tot_cp += d["cp"]
        tot_min += d["emin"]; tot_max += d["emax"]
        row_n += 1

    for col, val in enumerate([
        "TOTAL INVENTARIO TASADO", f"{tot_n} artículos",
        f"{tot_cp} con precio",
        f"$ {tot_min:,}", f"$ {tot_max:,}", ""
    ], 1):
        hdr(ws, row_n, col, val, bg=SLATE, size=10)
    ws.row_dimensions[row_n].height = 22
    row_n += 2

    reservados = sum(1 for i in items if i["reservado"])
    ws.merge_cells(f"A{row_n}:F{row_n}")
    c = ws[f"A{row_n}"]
    c.value = (
        f"Del total de {tot_n} artículos valorados:  "
        f"{reservados} han sido reservados por un heredero  ·  "
        f"{tot_n - reservados} disponibles para venta o partición.  "
        "Valores Sotheby's enero 2016 — carácter exclusivamente referencial."
    )
    c.font      = Font(name="Calibri", size=9, italic=True, color="555555")
    c.fill      = fill(GOLD_LITE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = thin_border()
    ws.row_dimensions[row_n].height = 40


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — DETALLE POR CATEGORÍA  (misma estructura que Detalle por Heredero)
# N° | Código | Descripción | Ref. Sotheby's | Pág. | Estimación | Estado
# ══════════════════════════════════════════════════════════════════════════════
def hoja_detalle(wb):
    ws = wb.create_sheet("Detalle por Categoría")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    for col, w in zip("ABCDEFG", [5, 10, 44, 14, 6, 20, 24]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, 7,
        "DETALLE COMPLETO — ARTÍCULOS CON VALORACIÓN SOTHEBY'S · Sucesión Pignatelli",
        bg=NAVY, size=14, height=32)
    merge_sub(ws, 2, 7,
        "Sotheby's International Realty · Londres · 18 de enero de 2016  |  Valores en USD",
        bg=SLATE, height=20)
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(
        ["N°", "CÓDIGO", "DESCRIPCIÓN DEL OBJETO",
         "REF.\nSOTHEBY'S", "PÁG.", "ESTIMACIÓN (USD)", "RESERVADO PARA"], 1):
        hdr(ws, 3, col, h, bg=GRAY_HDR, size=9)
    ws.row_dimensions[3].height = 26
    ws.auto_filter.ref = "A3:G3"

    row_n = 4
    num   = 0
    cat_actual = None

    for item in items:
        cat    = item["categoria"]
        colors = CAT_COLORS.get(cat, {"dark": "444444", "light": "F5F5F5"})

        # Banda de categoría — igual que banda de heredero en Reservas
        if cat != cat_actual:
            cat_count = sum(1 for i in items if i["categoria"] == cat)
            ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=7)
            c = ws[f"A{row_n}"]
            c.value     = f"  {cat.upper()}  —  {cat_count} artículo{'s' if cat_count>1 else ''}"
            c.font      = Font(name="Calibri", size=12, bold=True, color=WHITE)
            c.fill      = fill(colors["dark"])
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_n].height = 24
            row_n    += 1
            cat_actual = cat

        # Fila de artículo
        num += 1
        lo, hi  = parse_est(item["est"])
        idx_cat = sum(1 for i in items[:items.index(item)] if i["categoria"] == cat)
        bg      = colors["light"] if idx_cat % 2 == 0 else WHITE
        est     = item["est"] if item["est"] not in ["-", ""] else "—"
        reservado = item["reservado"] or "—"

        dat(ws, row_n, 1, num,                         bg=bg, center=True, fg="666666")
        dat(ws, row_n, 2, item["codigo"],              bg=bg, center=True, bold=True,
            fg=colors["dark"])
        dat(ws, row_n, 3, item["nombre"].capitalize(), bg=bg)
        dat(ws, row_n, 4, f"Ref. {item['ref']}",       bg=bg, center=True,
            bold=True, fg=GOLD)
        dat(ws, row_n, 5, item["pag"],                 bg=bg, center=True, fg="777777")
        dat(ws, row_n, 6, est,                         bg=bg, center=True,
            bold=bool(lo), fg="1F4E79" if lo else "AAAAAA", italic=(est == "—"))
        dat(ws, row_n, 7, reservado,                   bg=bg,
            italic=(reservado == "—"), fg="555555" if reservado == "—" else "1A1A2E")
        ws.row_dimensions[row_n].height = 20
        row_n += 1

    # Total final
    tot_lo = sum(parse_est(i["est"])[0] or 0 for i in items)
    tot_hi = sum(parse_est(i["est"])[1] or 0 for i in items)
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
    c = ws[f"A{row_n}"]
    c.value     = f"  TOTAL — {num} artículos con valoración Sotheby's"
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border()
    c2 = ws.cell(row=row_n, column=7, value=f"$ {tot_lo:,} – $ {tot_hi:,}")
    c2.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c2.fill      = fill(SLATE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border    = thin_border()
    ws.row_dimensions[row_n].height = 22


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — RESERVADOS vs DISPONIBLES
# ══════════════════════════════════════════════════════════════════════════════
def hoja_estado(wb):
    ws = wb.create_sheet("Reservados vs Disponibles")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    for col, w in zip("ABCDEFG", [5, 10, 44, 18, 14, 20, 24]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, 7,
        "ESTADO DE LOS ARTÍCULOS TASADOS — RESERVADOS VS. DISPONIBLES",
        bg=NAVY, size=14, height=32)
    merge_sub(ws, 2, 7,
        "De los 57 artículos valorados por Sotheby's: cuáles han sido reservados y cuáles están disponibles",
        bg=SLATE, height=20)
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(
        ["N°", "CÓDIGO", "DESCRIPCIÓN", "CATEGORÍA",
         "REF. SOTHEBY'S", "ESTIMACIÓN (USD)", "RESERVADO PARA"], 1):
        hdr(ws, 3, col, h, bg=GRAY_HDR, size=9)
    ws.row_dimensions[3].height = 24
    ws.auto_filter.ref = "A3:G3"

    row_n    = 4
    reservados  = [i for i in items if i["reservado"]]
    disponibles = [i for i in items if not i["reservado"]]

    for color_banda, lista, label in [
        ("C0392B", reservados,  "RESERVADOS"),
        ("27AE60", disponibles, "DISPONIBLES PARA VENTA O PARTICIÓN"),
    ]:
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=7)
        c = ws[f"A{row_n}"]
        c.value     = f"  {label}  —  {len(lista)} artículo{'s' if len(lista)>1 else ''}"
        c.font      = Font(name="Calibri", size=12, bold=True, color=WHITE)
        c.fill      = fill(color_banda)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_n].height = 24
        row_n += 1

        sec_min = sec_max = 0
        for idx, item in enumerate(lista):
            lo, hi = parse_est(item["est"])
            if lo: sec_min += lo; sec_max += hi
            bg  = "FEF9E7" if idx % 2 == 0 else WHITE
            est = item["est"] if item["est"] not in ["-", ""] else "—"

            dat(ws, row_n, 1, idx + 1,                     bg=bg, center=True, fg="888888")
            dat(ws, row_n, 2, item["codigo"],              bg=bg, center=True, bold=True, fg="1A1A2E")
            dat(ws, row_n, 3, item["nombre"].capitalize(), bg=bg)
            dat(ws, row_n, 4, item["categoria"],           bg=bg, fg="555555")
            dat(ws, row_n, 5, f"Ref. {item['ref']}",       bg=bg, center=True, bold=True, fg=GOLD)
            dat(ws, row_n, 6, est,                         bg=bg, center=True,
                bold=bool(lo), fg="1F4E79" if lo else "AAAAAA", italic=(est == "—"))
            dat(ws, row_n, 7,
                item["reservado"] if item["reservado"] else "Disponible",
                bg=bg, italic=not bool(item["reservado"]), fg="555555")
            ws.row_dimensions[row_n].height = 20
            row_n += 1

        # Subtotal sección
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
        cs = ws[f"A{row_n}"]
        cs.value     = f"  {len(lista)} artículo{'s' if len(lista)>1 else ''}"
        cs.font      = Font(name="Calibri", size=9, italic=True, color="555555")
        cs.fill      = fill("F0F4F8")
        cs.alignment = Alignment(horizontal="left", vertical="center")
        cs.border    = thin_border()
        c2 = ws.cell(row=row_n, column=7,
                     value=f"$ {sec_min:,} – $ {sec_max:,}" if sec_min else "—")
        c2.font      = Font(name="Calibri", size=9,
                           bold=bool(sec_min), color="1F4E79" if sec_min else "888888")
        c2.fill      = fill("F0F4F8")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border    = thin_border()
        ws.row_dimensions[row_n].height = 16
        row_n += 2


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    hoja_resumen(wb)
    hoja_detalle(wb)
    hoja_estado(wb)

    wb.properties.title   = "Inventario Sotheby's — Sucesión Pignatelli 2026"
    wb.properties.subject = "Artículos del Inventario Solaris con Valoración Sotheby's"
    wb.properties.creator = "Sistema Pignatelli"
    wb.save(OUT_PATH)

    tot_lo = sum(parse_est(i["est"])[0] or 0 for i in items)
    tot_hi = sum(parse_est(i["est"])[1] or 0 for i in items)
    reservados = sum(1 for i in items if i["reservado"])

    print(f"\nOK  {OUT_PATH.name}")
    print(f"    Artículos tasados : {len(items)}")
    print(f"    Reservados        : {reservados}  |  Disponibles: {len(items) - reservados}")
    print(f"    Rango total       : $ {tot_lo:,} – $ {tot_hi:,} USD")
    print(f"    Hojas             : Resumen | Detalle por Categoría | Reservados vs Disponibles")

if __name__ == "__main__":
    main()
