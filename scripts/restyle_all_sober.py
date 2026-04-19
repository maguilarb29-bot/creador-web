"""
Apply institutional palette to ALL project Excel files.
Normalizes alpha prefix (00/FF) for matching.
Status colors (available/reserved/sold) are kept distinguishable but muted.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import os

BASE = r"C:\Users\Alejandro\Documents\Proyecto Pignatelli"
FILES = [
    r"API_PG_DRIVE\Inventario Maestro\Inventario_Maestro_Solaris_Pignatelli_2026-04-14.xlsx",
    r"API_PG_DRIVE\Inventario Maestro\Reservas de Herederos\01_Documentos\02_Reservas Herederos Pignatelli.xlsx",
    r"API_PG_DRIVE\Inventario Maestro\Heirs Reservations\01_Documents\02_Heirs Reservations Pignatelli.xlsx",
    r"API_PG_DRIVE\Inventario Maestro\Sothebys\01_Documentos\02_Sothebys Inventario Completo Pignatelli.xlsx",
    r"Api_PG\Reservas_Herederos_Pignatelli_2026-04-13.xlsx",
    r"docs\Inventario_Maestro_Solaris_Pignatelli_2026-04-14.xlsx",
    r"docs\Precios_Para_Asignar_2026-04-14.xlsx",
    r"docs\Reservas_Herederos_Pignatelli_2026-04-13.xlsx",
    r"docs\Sothebys\Sothebys_Inventario_Completo_Pignatelli_2026-04-13.xlsx",
]

# ── Color maps (keyed by last 6 hex chars, case-insensitive) ──────────────────

# Fill backgrounds
FILL_MAP = {
    # ── Titles / dark headers → institutional charcoal ──
    "1A1A2E": "002C3143",
    "2C3E50": "002C3143",
    "1A5276": "002C3143",
    "263545": "002C3143",
    "1C2833": "002C3143",
    "5D3A00": "002C3143",   # Sotheby's brown title
    # ── Subtitles / info bars → medium charcoal ──
    "3D5166": "003E4E5F",
    "34495E": "003E4E5F",
    "7B5200": "003E4E5F",
    # ── Column headers → steel blue-gray ──
    "37474F": "00546E7A",
    "455A64": "00546E7A",
    "424242": "00546E7A",
    # ── Per-heir banners (all colors) → single muted olive ──
    "1F4E79": "004E5E50",
    "2E6B30": "004E5E50",
    "7B3F00": "004E5E50",
    "5B1F5E": "004E5E50",
    "1A4480": "004E5E50",
    # ── Status dark headers (muted but distinct) ──
    "1E8449": "003D6B50",   # green  → muted dark green  (available)
    "2E7D32": "003D6B50",
    "27AE60": "003D6B50",
    "558B2F": "003D6B50",
    "1B5E20": "003D6B50",
    "1A3A2A": "003D6B50",
    "00695C": "003A5560",   # teal
    "0D47A1": "003A5080",   # blue
    "1565C0": "003A5080",
    "C0392B": "006B3A3A",   # red    → muted burgundy     (reserved)
    "922B21": "006B3A3A",
    "880E4F": "005E3A4A",   # deep pink
    "E65100": "006B5030",   # orange → muted warm brown
    "7B1FA2": "004A4068",   # purple → muted indigo       (sold)
    "6A1B9A": "003E3660",
    "4A148C": "003E3660",
    "7D3C98": "004A4068",
    "5D4037": "003E4E5F",   # brown → charcoal-gray
    "B7770D": "003E4E5F",
    # ── Light tinted row fills → uniform near-white ──
    "D5F5E3": "00F2F3F1",
    "D6EAF8": "00F2F3F1",
    "E0F2F1": "00F2F3F1",
    "E3F2FD": "00F2F3F1",
    "E8DAEF": "00F2F3F1",
    "E8F5E9": "00F2F3F1",
    "ECEFF1": "00F2F3F1",
    "EEF2F7": "00F2F3F1",
    "F1F8E9": "00F2F3F1",
    "F3E5F5": "00F2F3F1",
    "F5F5F5": "00F2F3F1",
    "FADBD8": "00F2F3F1",
    "FCE4EC": "00F2F3F1",
    "FDEBD0": "00F2F3F1",
    "FFF3E0": "00F2F3F1",
    "FFF8E1": "00F2F3F1",
    "DCE9F5": "00F2F3F1",
    "DFF0DC": "00F2F3F1",
    "FDEEDE": "00F2F3F1",
    "F2E0F4": "00F2F3F1",
    "D9E8F6": "00F2F3F1",
    # ── Subtotal rows → light gray ──
    "F0F4F8": "00E7E9E7",
    "E8EAF6": "00E7E9E7",
    # ── Note / footer → near-white ──
    "FEF9E7": "00F7F7F7",
    "FFFDE7": "00F7F7F7",
    "FFF9E7": "00F7F7F7",
}

# Font colors
FONT_MAP = {
    # ── Per-heir accent → near-black ──
    "1F4E79": "001A1A1A",
    "2E6B30": "001A1A1A",
    "7B3F00": "001A1A1A",
    "5B1F5E": "001A1A1A",
    "1A4480": "001A1A1A",
    "1A3A5C": "001A1A1A",
    "0D47A1": "00374A64",
    "1565C0": "00374A64",
    "1A5276": "00374A64",
    "1A1A2E": "002C3143",
    # ── Status font colors → muted equivalents ──
    "1E8449": "003D6B50",
    "2E7D32": "003D6B50",
    "27AE60": "003D6B50",
    "558B2F": "003D6B50",
    "1B5E20": "003D6B50",
    "00695C": "003A5560",
    "C0392B": "006B3A3A",
    "922B21": "006B3A3A",
    "880E4F": "005E3A4A",
    "E65100": "006B5030",
    "7B1FA2": "004A4068",
    "6A1B9A": "003E3660",
    "4A148C": "003E3660",
    "7D3C98": "004A4068",
    "5D4037": "00546E7A",
    # ── Gold/amber → steel gray ──
    "B8860B": "005A6470",
    "B7770D": "005A6470",
    "CC8800": "005A6470",
    # ── Saturated column header darks → steel ──
    "37474F": "00546E7A",
    "424242": "00444444",
    # ── Muted/placeholder → normalize ──
    "BBBBBB": "00888888",
    "AAAAAA": "00888888",
    "CCCCCC": "00888888",
    "999999": "00888888",
    # ── Subtitle light text ──
    "D0D8E0": "00B0BEC5",
    "D0B0E0": "00B0BEC5",
    "D0D0D0": "00B0BEC5",
    "BBDEFB": "00B0BEC5",
    "FFE0A0": "00B0BEC5",
}

def rgb6(raw):
    """Extract last 6 hex chars, uppercase."""
    if not raw or len(raw) < 6:
        return None
    return raw[-6:].upper()

def is_blank(raw):
    if not raw:
        return True
    r = raw.upper()
    return r in ("00000000", "FFFFFFFF", "00FFFFFF", "FF000000", "00000000")

def restyle(path):
    if not os.path.exists(path):
        print(f"  SKIP (not found): {os.path.basename(path)}")
        return
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"  ERROR loading: {e}")
        return

    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                # Fill
                try:
                    fg = cell.fill.fgColor
                    if fg and fg.type == "rgb" and not is_blank(fg.rgb):
                        key = rgb6(fg.rgb)
                        if key and key in FILL_MAP:
                            cell.fill = PatternFill("solid", fgColor=FILL_MAP[key])
                            changed += 1
                except Exception:
                    pass
                # Font color
                try:
                    fc = cell.font.color
                    if fc and fc.type == "rgb" and not is_blank(fc.rgb):
                        key = rgb6(fc.rgb)
                        if key and key in FONT_MAP:
                            f = cell.font
                            cell.font = Font(
                                bold=f.bold, italic=f.italic, size=f.size,
                                color=FONT_MAP[key], name=f.name,
                                underline=f.underline, strike=f.strike
                            )
                            changed += 1
                except Exception:
                    pass

    try:
        wb.save(path)
        print(f"  OK  ({changed} changes): {os.path.basename(path)}")
    except PermissionError:
        print(f"  LOCKED (close it first): {os.path.basename(path)}")

print("Restyling all Excel files...\n")
for rel in FILES:
    restyle(os.path.join(BASE, rel))
print("\nDone.")
