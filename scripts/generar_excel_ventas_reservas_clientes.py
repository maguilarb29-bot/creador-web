"""
Excel de Ventas y Reservas de Clientes Externos — Solaris / Sucesion Pignatelli
Hoja 1: Ventas realizadas (por cliente)
Hoja 2: Reservas pendientes (por cliente)
"""
import json
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE     = Path(__file__).resolve().parent.parent
OUT_PATH = BASE / "docs" / "Ventas_Reservas_Clientes_2026-04-17.xlsx"

NAVY  = "1A1A2E"; SLATE = "2C3E50"; WHITE = "FFFFFF"

CLIENTE_COLORS = {
    "Rasmi":          {"dark": "0D47A1", "light": "E3F2FD"},
    "Nuria":          {"dark": "880E4F", "light": "FCE4EC"},
    "Alejandro":      {"dark": "E65100", "light": "FFF3E0"},
    "Pablo Brenes":   {"dark": "2E7D32", "light": "E8F5E9"},
    "Marielos Alpizar":{"dark":"4A148C", "light": "F3E5F5"},
    "Ana Martin":     {"dark": "00695C", "light": "E0F2F1"},
    "Luciana Paris":  {"dark": "F57F17", "light": "FFFDE7"},
    "Evelina Vargas": {"dark": "00838F", "light": "E0F7FA"},
}

HEREDEROS = {
    "Adriano Pignatelli","Diego Pignatelli","Fabrizia Pignatelli",
    "Margherita Pignatelli","Maria Cristina Smith",
}

def s():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def fill(h): return PatternFill("solid", fgColor=h)

def cell(ws, row, col, val, bg=WHITE, fg="1A1A1A", bold=False,
         italic=False, center=False, size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=bold, italic=italic, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border    = s()

def banda(ws, row, cols, texto, color_dark):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws[f"A{row}"]
    c.value     = texto
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill(color_dark)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20

def total_row(ws, row, cols, texto, extra_cells=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws[f"A{row}"]
    c.value     = texto
    c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = s()
    if extra_cells:
        for col, val in extra_cells:
            cx = ws.cell(row=row, column=col, value=val)
            cx.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
            cx.fill      = fill(SLATE)
            cx.alignment = Alignment(horizontal="center", vertical="center")
            cx.border    = s()
    ws.row_dimensions[row].height = 22


# ── Cargar datos ──────────────────────────────────────────────────────────────
with open(BASE / "Api_PG/data/solaris_catalogo.json", encoding="utf-8-sig") as f:
    catalogo = json.load(f)

articulos = [a for a in catalogo
             if a.get("tipoEstructural") in ("ARTICULO","SET","LOTE")]

clientes_externos = {a.get("reservadoPara")
                     for a in articulos
                     if a.get("reservadoPara") and a.get("reservadoPara") not in HEREDEROS}

vendidos_cli  = [a for a in articulos
                 if a.get("estado") == "Vendido"
                 and a.get("reservadoPara") not in HEREDEROS
                 and a.get("reservadoPara")]

reservas_cli  = [a for a in articulos
                 if a.get("estado") == "Reservado"
                 and a.get("reservadoPara") not in HEREDEROS
                 and a.get("reservadoPara")]

ventas_por_cli  = defaultdict(list)
reservas_por_cli= defaultdict(list)
for a in vendidos_cli:  ventas_por_cli[a["reservadoPara"]].append(a)
for a in reservas_cli:  reservas_por_cli[a["reservadoPara"]].append(a)

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — VENTAS A CLIENTES
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Ventas Clientes"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

COLS_V = 8
for col, w in zip("ABCDEFGH", [5, 10, 44, 20, 14, 16, 18, 14]):
    ws.column_dimensions[col].width = w

ws.merge_cells("A1:H1")
c = ws["A1"]
c.value     = "VENTAS A CLIENTES - Condominio Solaris / Sucesion Pignatelli"
c.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill      = fill("6A1B9A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:H2")
c = ws["A2"]
c.value     = "Ventas realizadas a compradores externos  |  Precios en USD y CRC  |  Abril 2026"
c.font      = Font(name="Calibri", size=9, italic=True, color=WHITE)
c.fill      = fill("4A148C")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

ws.merge_cells("A3:H3")
ws["A3"].fill = fill("EEF2F7")
ws.row_dimensions[3].height = 8

for col, h in enumerate(
    ["N°","CODIGO","DESCRIPCION","COMPRADOR","CATEGORIA","PRECIO USD","PRECIO CRC","FECHA VENTA"], 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font      = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill      = fill("37474F")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = s()
ws.row_dimensions[4].height = 24
ws.auto_filter.ref = "A4:H4"

row_n = 5; num = 0; tot_usd = 0.0; tot_crc = 0.0

for cli in sorted(ventas_por_cli.keys()):
    items  = ventas_por_cli[cli]
    colors = CLIENTE_COLORS.get(cli, {"dark":"444444","light":"F5F5F5"})
    banda(ws, row_n, COLS_V,
          f"  {cli.upper()}  -  {len(items)} articulo{'s' if len(items)>1 else ''}",
          colors["dark"])
    row_n += 1

    for idx, a in enumerate(items):
        num += 1
        bg  = colors["light"] if idx % 2 == 0 else WHITE
        usd = a.get("precioUSD", 0) or 0
        crc = a.get("precioColones", 0) or 0
        tot_usd += usd; tot_crc += crc

        cell(ws, row_n, 1, num,                                    bg=bg, center=True, fg="888888", size=9)
        cell(ws, row_n, 2, a["codigoItem"],                        bg=bg, center=True, bold=True, fg=colors["dark"])
        cell(ws, row_n, 3, a.get("nombreES","").capitalize()[:55], bg=bg, size=9)
        cell(ws, row_n, 4, cli,                                    bg=bg, bold=True, fg=colors["dark"], size=9)
        cell(ws, row_n, 5, a.get("categoria",""),                  bg=bg, fg="666666", size=8, center=True)
        cell(ws, row_n, 6, f"$ {usd:,.2f}" if usd else "",        bg=bg, center=True, bold=True, fg="1F4E79")
        cell(ws, row_n, 7, f"CRC {crc:,.2f}" if crc else "",      bg=bg, center=True, bold=True, fg="27AE60", size=9)
        cell(ws, row_n, 8, a.get("fechaVenta",""),                 bg=bg, center=True, fg="555555", size=9)
        ws.row_dimensions[row_n].height = 18
        row_n += 1

total_row(ws, row_n, 5, f"  TOTAL - {num} articulos vendidos",
          [(6, f"$ {tot_usd:,.2f}"), (7, f"CRC {tot_crc:,.2f}"), (8, "")])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — RESERVAS DE CLIENTES
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Reservas Clientes")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A5"

COLS_R = 6
for col, w in zip("ABCDEF", [5, 10, 48, 20, 14, 22]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value     = "RESERVAS DE CLIENTES - Condominio Solaris / Sucesion Pignatelli"
c.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill      = fill("0D47A1")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 34

ws2.merge_cells("A2:F2")
c = ws2["A2"]
c.value     = "Articulos apartados por compradores externos - pendientes de cierre"
c.font      = Font(name="Calibri", size=9, italic=True, color=WHITE)
c.fill      = fill("1565C0")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[2].height = 18

ws2.merge_cells("A3:F3")
ws2["A3"].fill = fill("EEF2F7")
ws2.row_dimensions[3].height = 8

for col, h in enumerate(
    ["N°","CODIGO","DESCRIPCION","CLIENTE","CATEGORIA","PRECIO SUGERIDO (USD)"], 1):
    c = ws2.cell(row=4, column=col, value=h)
    c.font      = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill      = fill("37474F")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = s()
ws2.row_dimensions[4].height = 24
ws2.auto_filter.ref = "A4:F4"

row_n = 5; num = 0

for cli in sorted(reservas_por_cli.keys()):
    items  = reservas_por_cli[cli]
    colors = CLIENTE_COLORS.get(cli, {"dark":"444444","light":"F5F5F5"})
    banda(ws2, row_n, COLS_R,
          f"  {cli.upper()}  -  {len(items)} articulo{'s' if len(items)>1 else ''}",
          colors["dark"])
    row_n += 1

    for idx, a in enumerate(items):
        num += 1
        bg     = colors["light"] if idx % 2 == 0 else WHITE
        precio = a.get("precioUSD")

        cell(ws2, row_n, 1, num,                                    bg=bg, center=True, fg="888888", size=9)
        cell(ws2, row_n, 2, a["codigoItem"],                        bg=bg, center=True, bold=True, fg=colors["dark"])
        cell(ws2, row_n, 3, a.get("nombreES","").capitalize()[:65], bg=bg, size=9)
        cell(ws2, row_n, 4, cli,                                    bg=bg, bold=True, fg=colors["dark"], size=9)
        cell(ws2, row_n, 5, a.get("categoria",""),                  bg=bg, fg="666666", size=8, center=True)
        cell(ws2, row_n, 6,
             f"$ {precio:,.0f}" if precio else "Sin precio",
             bg="E8F5E9" if not precio else bg,
             center=True, bold=bool(precio),
             fg="1B5E20" if precio else "999999", italic=not bool(precio))
        ws2.row_dimensions[row_n].height = 18
        row_n += 1

total_row(ws2, row_n, COLS_R,
          f"  TOTAL - {num} articulos reservados por clientes")


wb.properties.title = "Ventas y Reservas Clientes - Solaris 2026"
wb.save(OUT_PATH)

print(f"\nOK  {OUT_PATH.name}")
print(f"    Ventas  : {len(vendidos_cli)} articulos  USD {tot_usd:,.2f} / CRC {tot_crc:,.2f}")
print(f"    Reservas: {len(reservas_cli)} articulos")

if __name__ == "__main__":
    pass
