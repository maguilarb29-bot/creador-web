"""
Restyle Excel to institutional/legal palette:
- Dark charcoal titles (single color for all)
- Steel-gray column headers
- ONE muted olive-gray for all heir section banners
- White / very-light-gray alternating rows
- All text near-black (no blue, no gold, no per-heir colors)
- Subtle gray for Sotheby's refs and N/A markers
"""

import openpyxl
import shutil
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SRC = r"C:\Users\Alejandro\Documents\Proyecto Pignatelli\API_PG_DRIVE\Inventario Maestro\Heirs Reservations\01_Documents\02_Heirs Reservations Pignatelli.xlsx"
DST = SRC  # overwrite in place

# ── Palette ──────────────────────────────────────────────────────────────────
C_TITLE       = "FF2C3143"   # dark charcoal-navy  — main title rows
C_SUBTITLE    = "FF3E4E5F"   # slightly lighter    — subtitle / info rows
C_COL_HEADER  = "FF546E7A"   # steel blue-gray     — column header rows
C_HEIR_BANNER = "FF4E5E50"   # muted olive-gray    — ALL heir section banners (single color)
C_ROW_ALT     = "FFF2F3F1"   # near-white warm     — alternating data rows
C_ROW_WHITE   = "FFFFFFFF"   # pure white          — data rows
C_SUBTOTAL    = "FFE7E9E7"   # light gray          — "X items reserved" rows
C_TOTAL       = "FF2C3143"   # same as title       — total row
C_NOTE        = "FFF7F7F7"   # near-white          — legal note / footer

TXT_WHITE  = "FFFFFFFF"
TXT_BLACK  = "FF1A1A1A"
TXT_MID    = "FF444444"
TXT_MUTED  = "FF888888"
TXT_HEADER = "FF3A4A3A"   # dark olive for subtotals
TXT_REF    = "FF5A6470"   # steel gray for Sotheby's refs (replaces gold)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=TXT_BLACK, size=None, italic=False):
    kw = dict(bold=bold, color=color)
    if size:  kw["size"] = size
    if italic: kw["italic"] = italic
    return Font(**kw)

def thin_border():
    s = Side(style="thin", color="FFD0D4D0")
    return Border(bottom=s)

def no_border():
    return Border()

# ── Load workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(SRC)

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET: Summary
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb["Summary"]

# Row 1 — main title
for cell in ws[1]:
    cell.fill = fill(C_TITLE)
    cell.font = font(bold=True, color=TXT_WHITE, size=12)

# Row 2 — subtitle
for cell in ws[2]:
    cell.fill = fill(C_SUBTITLE)
    cell.font = font(bold=False, color=TXT_WHITE)

# Row 3 — valuation note
for cell in ws[3]:
    cell.fill = fill(C_SUBTITLE)
    cell.font = font(bold=False, color="FFB0BEC5")  # light steel, subtle

# Row 4 — empty spacer, clear it
for cell in ws[4]:
    cell.fill = fill(C_ROW_WHITE)

# Row 5 — column headers
for cell in ws[5]:
    cell.fill = fill(C_COL_HEADER)
    cell.font = font(bold=True, color=TXT_WHITE)

# Rows 6-10 — heir data rows: all same light alternating, black text
for row_idx, row in enumerate(ws.iter_rows(min_row=6, max_row=10)):
    bg = C_ROW_ALT if row_idx % 2 == 0 else C_ROW_WHITE
    for cell in row:
        cell.fill = fill(bg)
        if cell.column == 1:  # heir name
            cell.font = font(bold=True, color=TXT_BLACK)
        else:
            cell.font = font(bold=True, color=TXT_BLACK)

# Row 11 — TOTAL
for cell in ws[11]:
    cell.fill = fill(C_TOTAL)
    cell.font = font(bold=True, color=TXT_WHITE)

# Row 13 — legal note (skip row 12 spacer)
for cell in ws[12]:
    cell.fill = fill(C_ROW_WHITE)
for cell in ws[13]:
    cell.fill = fill(C_NOTE)
    cell.font = font(bold=False, color="FF555555", italic=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET: Detail by Heir
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb["Detail by Heir"]

def style_col_headers(row_cells):
    for cell in row_cells:
        cell.fill = fill(C_COL_HEADER)
        cell.font = font(bold=True, color=TXT_WHITE)

def style_data_row(row_cells, alt=False):
    bg = C_ROW_ALT if alt else C_ROW_WHITE
    for cell in row_cells:
        if cell.value is None:
            cell.fill = fill(bg)
            continue
        v = str(cell.value) if cell.value else ""
        cell.fill = fill(bg)
        col = cell.column
        if col == 1:   # No.
            cell.font = font(bold=False, color=TXT_MUTED)
        elif col == 2: # CODE
            cell.font = font(bold=True, color=TXT_BLACK)
        elif col == 3: # DESCRIPTION
            cell.font = font(bold=False, color=TXT_BLACK)
        elif col == 4: # CATEGORY
            cell.font = font(bold=False, color=TXT_MID)
        elif col == 5: # PHOTOS
            cell.font = font(bold=False, color=TXT_MUTED)
        elif col == 6: # SOTHEBY'S REF
            if v in ("—", "\u2014", ""):
                cell.font = font(bold=False, color=TXT_MUTED)
            else:
                cell.font = font(bold=False, color=TXT_REF)
        elif col == 7: # ESTIMATE
            if v in ("—", "\u2014", ""):
                cell.font = font(bold=False, color=TXT_MUTED)
            else:
                cell.font = font(bold=True, color=TXT_BLACK)

# Row 1 — title
for cell in ws[1]:
    cell.fill = fill(C_TITLE)
    cell.font = font(bold=True, color=TXT_WHITE, size=12)

# Row 2 — subtitle
for cell in ws[2]:
    cell.fill = fill(C_SUBTITLE)
    cell.font = font(bold=False, color=TXT_WHITE)

# Row 3 — spacer
for cell in ws[3]:
    cell.fill = fill(C_ROW_WHITE)

# Iterate all remaining rows and classify them
alt_counter = 0
for row in ws.iter_rows(min_row=4):
    first_val = row[0].value
    if first_val is None:
        # spacer
        for cell in row:
            cell.fill = fill(C_ROW_WHITE)
            cell.font = font()
        alt_counter = 0
        continue

    s = str(first_val).strip()

    # Heir section banner (starts with spaces + ALL CAPS name)
    if s.upper() == s and len(s) > 3 and not s.startswith("No") and not any(c.isdigit() for c in s[:2]):
        for cell in row:
            cell.fill = fill(C_HEIR_BANNER)
            cell.font = font(bold=True, color=TXT_WHITE)
        alt_counter = 0

    # Column header row
    elif s == "No.":
        style_col_headers(row)

    # Subtotal row ("X items reserved")
    elif "items reserved" in s:
        for cell in row:
            cell.fill = fill(C_SUBTOTAL)
            if cell.value:
                cell.font = font(bold=False, color="FF4A5568", italic=True)
            else:
                cell.font = font()

    # Data rows
    elif s[:2].strip().isdigit() or (len(s) > 0 and s[0].isdigit()):
        style_data_row(row, alt=(alt_counter % 2 == 0))
        alt_counter += 1

    else:
        for cell in row:
            cell.fill = fill(C_ROW_WHITE)
            cell.font = font()

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET: Sotheby's Valuation
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb["Sotheby's Valuation"]

# Row 1 — title (was brown, now charcoal)
for cell in ws[1]:
    cell.fill = fill(C_TITLE)
    cell.font = font(bold=True, color=TXT_WHITE, size=12)

# Row 2 — subtitle
for cell in ws[2]:
    cell.fill = fill(C_SUBTITLE)
    cell.font = font(bold=False, color="FFB0BEC5")

# Row 3 — spacer
for cell in ws[3]:
    cell.fill = fill(C_ROW_WHITE)

# Row 4 — column headers (was brown, now steel-gray)
for cell in ws[4]:
    cell.fill = fill(C_COL_HEADER)
    cell.font = font(bold=True, color=TXT_WHITE)

# Data rows 5-22
alt_counter = 0
for row in ws.iter_rows(min_row=5, max_row=22):
    bg = C_ROW_ALT if alt_counter % 2 == 0 else C_ROW_WHITE
    for cell in row:
        cell.fill = fill(bg)
        col = cell.column
        v = str(cell.value) if cell.value else ""
        if col == 1:   # Sotheby's ref
            cell.font = font(bold=False, color=TXT_REF)
        elif col == 2: # page
            cell.font = font(bold=False, color=TXT_MUTED)
        elif col == 3: # code
            cell.font = font(bold=True, color=TXT_BLACK)
        elif col == 4: # description
            cell.font = font(bold=False, color=TXT_BLACK)
        elif col == 5: # heir
            cell.font = font(bold=False, color=TXT_MID)
        elif col == 6: # estimate
            if v in ("—", "\u2014", ""):
                cell.font = font(bold=False, color=TXT_MUTED)
            else:
                cell.font = font(bold=True, color=TXT_BLACK)
    alt_counter += 1

# Row 23 — TOTAL APPRAISED
for cell in ws[23]:
    cell.fill = fill(C_TOTAL)
    cell.font = font(bold=True, color=TXT_WHITE)

# Row 24 — spacer
if ws.max_row >= 24:
    for cell in ws[24]:
        cell.fill = fill(C_ROW_WHITE)

# Row 25 — footer note
if ws.max_row >= 25:
    for cell in ws[25]:
        cell.fill = fill(C_NOTE)
        cell.font = font(bold=False, color="FF555555", italic=True)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(DST)
print("Restyled and saved OK:", DST)
