import openpyxl, shutil

SRC = r"C:\Users\Alejandro\Documents\Proyecto Pignatelli\API_PG_DRIVE\Inventario Maestro\Reservas de Herederos\01_Documentos\02_Reservas Herederos Pignatelli.xlsx"
DST = r"C:\Users\Alejandro\Documents\Proyecto Pignatelli\API_PG_DRIVE\Inventario Maestro\Heirs Reservations\01_Documents\02_Heirs Reservations Pignatelli.xlsx"

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)

SHEET_NAMES = {
    "Resumen": "Summary",
    "Detalle por Heredero": "Detail by Heir",
    "Valoraci\xf3n Sotheby's": "Sotheby's Valuation",
}

FULL = {
    # Titles with em-dash (U+2014)
    "SUCESI\xd3N PIGNATELLI \u2014 CONDOMINIO SOLARIS":
        "PIGNATELLI ESTATE \u2014 CONDOMINIO SOLARIS",
    "Reservas de Herederos sobre Inventario Solaris \u2014 Abril 2026":
        "Heirs' Reservations on Solaris Inventory \u2014 April 2026",
    "DETALLE DE RESERVAS POR HEREDERO \u2014 SUCESI\xd3N PIGNATELLI":
        "RESERVATION DETAILS BY HEIR \u2014 PIGNATELLI ESTATE",
    "VALORACI\xd3N SOTHEBY'S \u2014 \xcdTEMS RESERVADOS":
        "SOTHEBY'S VALUATION \u2014 RESERVED ITEMS",
    "  TOTAL TASADO  \u2014  18 \xedtems con referencia Sotheby's":
        "  TOTAL APPRAISED  \u2014  18 items with Sotheby's reference",

    # Sotheby's header with middle-dot (U+00B7)
    "Sotheby's International Realty \xb7 Londres \xb7 18 de enero de 2016  |  Valores en USD":
        "Sotheby's International Realty \xb7 London \xb7 January 18, 2016  |  Values in USD",

    # Valuation line (no special dash)
    "Valoraciones: Sotheby's International Realty, Londres, 18 de enero de 2016  (USD)":
        "Valuations: Sotheby's International Realty, London, January 18, 2016  (USD)",
    "Valoraci\xf3n Sotheby's 2016  |  Inventario fotogr\xe1fico oficial Condominio Solaris":
        "Sotheby's Valuation 2016  |  Official Photographic Inventory Condominio Solaris",

    # Column headers
    "HEREDERO": "HEIR",
    "\xcdTEMS\nRESERVADOS": "RESERVED\nITEMS",
    "\xcdTEMS CON\nTASACI\xd3N": "ITEMS WITH\nVALUATION",
    "ESTIMADO\nM\xcdNIMO (USD)": "MINIMUM\nESTIMATE (USD)",
    "ESTIMADO\nM\xc1XIMO (USD)": "MAXIMUM\nESTIMATE (USD)",
    "TOTAL": "TOTAL",
    "N\xb0": "No.",
    "C\xd3DIGO": "CODE",
    "DESCRIPCI\xd3N": "DESCRIPTION",
    "CATEGOR\xcdA": "CATEGORY",
    "FOTOS": "PHOTOS",
    "REF.\nSOTHEBY'S": "SOTHEBY'S\nREF.",
    "ESTIMACI\xd3N\n(USD)": "ESTIMATE\n(USD)",
    "ESTIMACI\xd3N (USD)": "ESTIMATE (USD)",
    "DESCRIPCI\xd3N DEL OBJETO": "OBJECT DESCRIPTION",
    "P\xc1G.": "PAGE",

    # Per-heir summary rows (Rango uses en-dash U+2013)
    "  10 \xedtems reservados": "  10 items reserved",
    "  9 \xedtems reservados": "  9 items reserved",
    "  4 \xedtems reservados": "  4 items reserved",
    "  2 \xedtems reservados": "  2 items reserved",
    "Rango: $ 12,880 \u2013 $ 17,780": "Range: $ 12,880 \u2013 $ 17,780",
    "Rango: $ 26,260 \u2013 $ 35,640": "Range: $ 26,260 \u2013 $ 35,640",
    "Rango: $ 14,420 \u2013 $ 21,280": "Range: $ 14,420 \u2013 $ 21,280",
    "Rango: $ 19,600 \u2013 $ 29,400": "Range: $ 19,600 \u2013 $ 29,400",
    "Rango: $ 8,716 \u2013 $ 11,516": "Range: $ 8,716 \u2013 $ 11,516",

    # Categories
    "Cristaleria": "Glassware",
    "Arte en papel": "Works on Paper",
    "Plateria": "Silverware",
    "Decorativos": "Decorative Items",
    "Utensilios": "Utensils",
    "Joyas": "Jewelry",
    "Muebles": "Furniture",

    # Item descriptions
    "Cristaleria tallada roja y dorada": "Red and Gold Cut Glassware",
    "Pintura de pareja paseando": "Painting of Couple Strolling",
    "Figuras en terrazas chinas del siglo xviii": "Figures on 18th-Century Chinese Terraces",
    "Pintura de dos leones": "Painting of Two Lions",
    "Tres grabados historicos": "Three Historical Engravings",
    "Juego de cubiertos de plata y asta para carne": "Silver and Bone-Handle Carving Set",
    "Jarra de vidrio soplado": "Blown Glass Pitcher",
    "Lote de libros": "Lot of Books",
    "Cristaleria de borde dorado": "Gold-Rimmed Glassware",
    "Caballo tang de ceramica": "Ceramic Tang Horse",
    "Anillos y broches de fantasia con gemas variadas": "Fantasy Rings and Brooches with Assorted Gems",
    "Plateri clasica con bandejas y cuencos decorados": "Classic Silverware with Decorated Trays and Bowls",
    "Cubiertos de plata en bandeja de bambu": "Silver Cutlery on Bamboo Tray",
    "Busto escultorico de cabeza masculina texturizada": "Sculptural Bust of Textured Male Head",
    "Servilleteros de plata y figuras de animales en estuche": "Silver Napkin Rings and Animal Figures in Case",
    "Cuencos plateados estilo imperio": "Empire-Style Silver Bowls",
    "Cubiertos dorados en estuche de lujo": "Gold Cutlery in Luxury Case",
    "Par de salseras francesas de plata": "Pair of French Silver Sauce Boats",
    "Reloj cartel luis xv bronce dorado saint german": "Louis XV Cartel Clock, Gilded Bronze, Saint-Germain",
    "Cubiertos dorados vintage en estuche de presentacion": "Vintage Gold Cutlery in Presentation Case",
    "Estuche de almacenamiento vintage de cuero y madera": "Vintage Leather and Wood Storage Case",
    "Comoda luis xv con marqueteria y bronces dorados": "Louis XV Commode with Marquetry and Gilded Bronzes",
    "Consola dorada con marmol fior di pesco": "Gilded Console with Fior di Pesco Marble",
    "Mesa de centro china lacada": "Chinese Lacquered Coffee Table",
    "Comoda bombe estilo luis xv con marqueteria y bronces": "Louis XV-Style Bombe Commode with Marquetry and Bronzes",

    # Misc
    "Ref. (Sin codigo)": "Ref. (No code)",
    "$ 81,876 \u2013 $ 115,616": "$ 81,876 \u2013 $ 115,616",
}

LEGAL_FRAGMENT = "reservas consignadas en este documento"
LEGAL_EN = (
    "LEGAL NOTE: The reservations recorded in this document represent the declared intention "
    "of each heir regarding the assets of the Condominio Solaris estate and do not constitute "
    "a definitive assignment until the partition process is formally concluded before the competent "
    "legal authorities. The valuations correspond to the Sotheby's International Realty appraisal "
    "dated January 18, 2016, and are for reference purposes only. Items without a Sotheby's "
    "reference were not included in that appraisal."
)

FOOTER_FRAGMENT = "tems reservados sin referencia"
FOOTER_EN = (
    "  Items reserved without reference in Sotheby's 2016 catalogue (9): "
    "162AC, 22A, 318A, 319A, 256A, 258A, 259A, 334A, 314A  \u2014  "
    "These items were not included in the 2016 appraisal or could not be identified "
    "with certainty in the catalogue (includes lots of books, costume jewellery and uncatalogued glassware)."
)

def translate(val):
    if not isinstance(val, str):
        return val
    # Try exact match first (preserves leading spaces in keys)
    if val in FULL:
        return FULL[val]
    # Try stripped match
    stripped = val.strip()
    if stripped in FULL:
        return FULL[stripped]
    if LEGAL_FRAGMENT in val:
        return LEGAL_EN
    if FOOTER_FRAGMENT in val:
        return FOOTER_EN
    return val

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.value = translate(cell.value)

for old, new in SHEET_NAMES.items():
    if old in wb.sheetnames:
        wb[old].title = new

wb.save(DST)
print("Done. Sheets:", wb.sheetnames)
