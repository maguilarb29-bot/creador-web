"""
Genera un Excel de trabajo para asignación de precios.
Dos hojas:
  1. Sin Precio (369 artículos) — para llenar
  2. Con Precio (149 artículos) — para revisar y corregir
"""

import json, re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE     = Path(__file__).resolve().parent.parent
OUT_PATH = BASE / "docs" / "Precios_Para_Asignar_2026-04-14.xlsx"

NAVY    = "1A1A2E"; SLATE = "2C3E50"; GRAY_HDR = "37474F"
WHITE   = "FFFFFF"; GOLD  = "B8860B"

CAT_COLORS = {
    "Arte en papel":    {"dark": "5D4037", "light": "FFF8E1"},
    "Ceramica":         {"dark": "1565C0", "light": "E3F2FD"},
    "Cristaleria":      {"dark": "00695C", "light": "E0F2F1"},
    "Decorativos":      {"dark": "4A148C", "light": "F3E5F5"},
    "Electrodomesticos":{"dark": "424242", "light": "F5F5F5"},
    "Insumos Medicos":  {"dark": "558B2F", "light": "F1F8E9"},
    "Joyas":            {"dark": "880E4F", "light": "FCE4EC"},
    "Muebles":          {"dark": "2E7D32", "light": "E8F5E9"},
    "Plateria":         {"dark": "37474F", "light": "ECEFF1"},
    "Utensilios":       {"dark": "E65100", "light": "FFF3E0"},
}

def side():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(h): return PatternFill("solid", fgColor=h)

def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, size=9):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=True, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = side()

def dat(ws, row, col, val, bg=WHITE, fg="1A1A1A", bold=False, italic=False,
        center=False, size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=size, bold=bold, italic=italic, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border    = side()

def merge_title(ws, row, cols, val, bg=NAVY, size=13, height=34):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=val)
    c.font      = Font(name="Calibri", size=size, bold=True, color=WHITE)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height

# ── Cargar datos ───────────────────────────────────────────────────────────────
with open(BASE / "Api_PG/data/solaris_catalogo.json", encoding="utf-8-sig") as f:
    catalogo = json.load(f)

import csv as _csv
soth_lookup = {}
with open(BASE / "docs/sothebys_maestro_2026-04-13.csv", newline="", encoding="utf-8-sig") as f:
    for row in _csv.DictReader(f):
        est = row.get("estimacion","").strip()
        for parte in row.get("codigo_actual","").split(";"):
            cod = re.sub(r"-\d+$","",parte.strip()).strip()
            if cod and est and est != "-":
                soth_lookup[cod] = est

articulos = [a for a in catalogo if a.get("tipoEstructural") in ("ARTICULO","SET","LOTE")]
articulos.sort(key=lambda x: (x.get("categoria",""), x.get("codigoItem","")))

sin_precio = [a for a in articulos if not a.get("precioUSD") and a.get("estado") != "Vendido" and not a.get("reservadoPara")]
con_precio = [a for a in articulos if a.get("precioUSD") and a.get("estado") != "Vendido"]

# ══════════════════════════════════════════════════════════════════════════════
def build_hoja(wb, title, items, sheet_name, fill_yellow=True):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    COLS = 7
    for col, w in zip("ABCDEFG", [5, 10, 46, 14, 20, 18, 26]):
        ws.column_dimensions[col].width = w

    merge_title(ws, 1, COLS, title, bg=NAVY if not fill_yellow else "1A3A2A")
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value     = "COLUMNA PRECIO SUGERIDO (USD): ingresa el valor sin símbolo de dólar  ·  Ej: 150"
    c.font      = Font(name="Calibri", size=9, italic=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(
        ["N°","CÓDIGO","DESCRIPCIÓN","CATEGORÍA",
         "VALORACIÓN SOTHEBY'S","PRECIO SUGERIDO (USD)","NOTAS"], 1):
        bg = GRAY_HDR if col != 6 else "1B5E20"  # verde oscuro para la col a llenar
        hdr(ws, 4, col, h, bg=bg)
    ws.row_dimensions[4].height = 26
    ws.auto_filter.ref = "A4:G4"

    row_n = 5
    cat_actual = None
    num = 0

    by_cat = defaultdict(list)
    for a in items:
        by_cat[a["categoria"]].append(a)

    for cat in sorted(by_cat.keys()):
        cat_items = by_cat[cat]
        colors = CAT_COLORS.get(cat, {"dark":"444444","light":"F5F5F5"})

        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
        c = ws[f"A{row_n}"]
        c.value     = f"  {cat.upper()}  -  {len(cat_items)} articulo{'s' if len(cat_items)>1 else ''}"
        c.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill      = fill(colors["dark"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_n].height = 20
        row_n += 1

        for idx, art in enumerate(cat_items):
            num += 1
            bg      = colors["light"] if idx % 2 == 0 else WHITE
            cod     = art["codigoItem"]
            soth    = soth_lookup.get(cod, "—")
            precio  = art.get("precioUSD")
            precio_v = precio if precio else ""  # vacío si no tiene — editable

            # Col precio: fondo verde claro si vacío (a llenar), normal si tiene
            bg_precio = "E8F5E9" if not precio else "FFFDE7"

            dat(ws, row_n, 1, num,                                  bg=bg, center=True, fg="888888", size=9)
            dat(ws, row_n, 2, cod,                                  bg=bg, center=True, bold=True, fg=colors["dark"])
            dat(ws, row_n, 3, art.get("nombreES","").capitalize()[:60], bg=bg, size=9)
            dat(ws, row_n, 4, cat,                                  bg=bg, fg="666666", size=8)
            dat(ws, row_n, 5, soth if soth != "—" else "",         bg=bg, center=True,
                bold=bool(soth and soth != "—"), fg=GOLD if soth != "—" else "CCCCCC", size=9)
            # Celda de precio — editable
            c_p = ws.cell(row=row_n, column=6, value=precio_v)
            c_p.font      = Font(name="Calibri", size=10, bold=bool(precio), color="1B5E20")
            c_p.fill      = fill(bg_precio)
            c_p.alignment = Alignment(horizontal="center", vertical="center")
            c_p.border    = side()
            dat(ws, row_n, 7, art.get("notas","") or "",           bg=bg, fg="888888", italic=True, size=8)
            ws.row_dimensions[row_n].height = 18
            row_n += 1

    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=COLS)
    c = ws[f"A{row_n}"]
    c.value     = f"  {num} artículos en esta hoja"
    c.font      = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill      = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = side()
    ws.row_dimensions[row_n].height = 20


def main():
    wb = Workbook()
    # Eliminar hoja default
    wb.remove(wb.active)

    build_hoja(wb,
        f"SIN PRECIO — {len(sin_precio)} artículos para valorar",
        sin_precio, "Sin Precio", fill_yellow=True)

    build_hoja(wb,
        f"CON PRECIO — {len(con_precio)} artículos para revisar",
        con_precio, "Con Precio", fill_yellow=False)

    wb.properties.title = "Precios para Asignar — Inventario Solaris 2026"
    wb.save(OUT_PATH)

    print(f"\nOK  {OUT_PATH.name}")
    print(f"    Sin precio : {len(sin_precio)} artículos  (col F en verde — llenar)")
    print(f"    Con precio : {len(con_precio)} artículos  (col F en amarillo — revisar)")

if __name__ == "__main__":
    main()
