import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\Alejandro\Documents\Proyecto Pignatelli\API_PG_DRIVE\Inventario Maestro\Reservas de Herederos\01_Documentos\02_Reservas Herederos Pignatelli.xlsx")
seen = set()
for shname in wb.sheetnames:
    ws = wb[shname]
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v and isinstance(v, str) and v not in seen:
                seen.add(v)
                esc = v.encode('unicode_escape').decode('ascii')
                has_nonascii = any(ord(c) > 127 for c in v)
                if has_nonascii:
                    print(esc[:150])
